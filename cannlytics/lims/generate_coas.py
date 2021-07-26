"""
Generate CoAs | Cannlytics

Author: Keegan Skeate <keegan@cannlytics.com>  
Created: 7/22/2021  
Updated: 7/23/2021  
License: MIT License <https://opensource.org/licenses/MIT>
"""
# Standard packages
import argparse
from ast import literal_eval
import os
from pathlib import Path
from re import sub, findall
from shutil import copyfile

# External packages
try:
    import openpyxl
    from openpyxl.drawing.image import Image
    import pandas as pd
    # import qrcode
    import win32com.client
    import xlwings
    from xlwings.utils import rgb_to_int
except:
    pass # FIXME: Docs can't import.


def calculate_results(sample_data, analysis, mass, dilution_factor=40, correction_factor=10000):
    """Calculate percentage results given raw results,
    dilution factor, and analysis type.
    Args:
        sample_data (dict): A dictionary of sample data.
        analysis (str): An analysis to calculate results for the analysis's analytes.
        mass (float): The recorded sample mass.
        dilution_factor (float): The dilution factor for the sample.
        correction_factor (float): A factor used to appropriately scale values to percentage.
    Returns:
        (dict): An updated dictionary of sample data.
    """
    analytes = get_analytes(analysis)
    for analyte in analytes:
        try:
            raw_value = float(sample_data[analyte])
            sample_data[analyte] = ((raw_value * dilution_factor) / mass) / correction_factor
        except ValueError:
            continue
    return sample_data


def create_coa_pdfs(render_file, ws_index_list, output_file, tight=False):
    """Generate PDFs for rendred CoAs.
    Args:
        render_file (str): The path of the rendred workbook.
        ws_index_list (list): A list of the worksheet indexes to include in the PDF.
        output_file (str): The name of the output PDF (expected .pdf extension).
        tight (bool): Optional, default False, choice to scale all pages.
    Returns:
        (COM Object): A Microsoft Excel Client to alter later.
    """
    client = win32com.client.Dispatch('Excel.Application')
    client.Visible = False
    client.ScreenUpdating = False
    client.DisplayAlerts = False
    client.EnableEvents = False
    dir_path = os.path.dirname(os.path.realpath(__file__))
    input_file = os.path.join(dir_path, render_file)
    pdf_file = os.path.join(dir_path, output_file)
    workbook = client.Workbooks.Open(input_file)
    if tight:
        print_area = 'A1:G50'
        for index in ws_index_list:
            worksheet = workbook.Worksheets[index]
            worksheet.PageSetup.Zoom = False
            worksheet.PageSetup.FitToPagesTall = 1
            worksheet.PageSetup.FitToPagesWide = 1
            worksheet.PageSetup.PrintArea = print_area
    try:
        workbook.WorkSheets(ws_index_list).Select()
    except:
        workbook.Worksheets(ws_index_list).Select()
    workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_file)
    workbook.Close(False)
    return client


def fill_jinja_references(workbook, data):
    """Fill-in Jinga-style references, iterating over all
    worksheet pages to fill-in all occurrences of each reference 1 by 1.
    Args:
        workbook (Workbook): A Workbook object.
        data (dict): A dictionary of context data.
    """
    context = { snake_case(key): values for key, values in data.items() }
    sheets = workbook.worksheets
    for sheet in sheets:
        refs = get_jinja_references(sheet)
        for key, cells in refs.items():
            clean_key = snake_case(key)
            value = context.get(clean_key, 0)
            for cell in cells:
                if value:
                    sheet[cell] = str(value)
                else:
                    sheet[cell] = 'MISSING'


def get_jinja_references(sheet):
    """Get Jinja-style references in an Excel worksheet.
    Args:
        sheet (Worksheet): A worksheet to get references from.
    Returns:
        (dict): A dictionary of variables to find references.
    """
    refs = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            try:
                if value.startswith('{{'):
                    ref = cell.column_letter + str(cell.row)
                    variable = value.replace('{{', '').replace('}}', '').strip()
                    existing_refs = refs.get(variable, [])
                    refs[variable] = [*existing_refs, *[ref]]
                elif value.startswith('{% for'): # Optional: Handle tables
                    return NotImplementedError
            except AttributeError:
                continue
    return refs


def get_worksheet_data(sheet, headers):
    """Get the data of a worksheet.
    Args:
        sheet (Worksheet): An openpyx; Excel file object.
        headres (list): A list of headers to map the values.
    Returns:
        list(dict): A list of dictionaries.
    """
    data = []
    for row in sheet.iter_rows():
        values = {}
        for key, cell in zip(headers, row):
            values[key] = cell.value
        data.append(values)
    return data


def get_worksheet_data_block(sheet, coords, expand=None):
    """Get a data block.
    Args:
        sheet (Sheet): The worksheet containing the data block.
        coords (str): The inclusive coordinates of the data block.
        expand (str): Optionally expand the range of values.
    Returns
        (dict): A dictionary of the data in the data block.
    """
    data = {}
    values = sheet.range(coords).options(expand=expand).value
    for item in values:
        key = snake_case(item[0])
        value = item[1]
        data[key] = value
    return data


def get_worksheet_headers(sheet):
    """Get the headres of a worksheet.
    Args:
        sheet (Worksheet): An openpyx; Excel file object.
    Returns:
        headers (list): A list of header strings.
    """
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    return headers


def get_worksheet_indexes(wb, output_pages):
    """Get the indexes for a list of sheet names in a given workbook.
    Args:
        wb (Workbook): The workbook at hand.
        pages (list): A list of pages to find indexes.
    """
    ws_index_list = []
    for page in output_pages:
        index = wb.worksheets.index(wb[page])
        ws_index_list.append(index + 1)
    return ws_index_list


def get_analytes(analysis, limits_file='analytes.xlsx', key='import_key'):
    """Get all analytes for a given analysis."""
    analytes = []
    workbook = openpyxl.load_workbook(limits_file, data_only=True)
    analyte_data = read_worksheet(workbook, sheetname='analytes')
    for analyte in analyte_data:
        if analyte['analysis_key'] == analysis:
            analytes.append(analyte[key])
    return analytes


def get_analyte_limits(limits_file='analytes.xlsx'):
    """Get analyte limits."""
    workbook = openpyxl.load_workbook(limits_file, data_only=True)
    analyte_data = read_worksheet(workbook, sheetname='analytes')
    limits = {}
    for analyte in analyte_data:
        key = analyte['import_key']
        limits[key + '_loq'] = analyte['loq']
        limits[key + '_limit'] = analyte['limit']
    return limits


def generate_coas(
        import_files,
        output_pages,
        coa_template='./coa_template.xlsm',
        # render_file='./CoAs/coa_render.xlsm',
        limits={}
):
    """Generate certificates of analysis.
    Args:
        import_files (list): A list of files to import.
        output_pages (list): A list of pages to include in the PDF.
        coa_template (str): The path of the CoA Template.
        limits (dict): A dictionary of limits and LOQ for analytes.
    """

    # Create CoA folder if one does not exist.
    Path('CoAs').mkdir(parents=True, exist_ok=True)
    dir_path = os.path.dirname(os.path.realpath(__file__))

    # Create a copy of the template.
    abs_coa_template = os.path.join(dir_path, coa_template)
    coa_template_copy = abs_coa_template.replace('.xlsm', '_copy.xlsm')
    copyfile(abs_coa_template, coa_template_copy)

    # Iterate over all import files.
    for import_file in import_files:

        # Get all sample, results, client, etc. data.
        # abs_import_file = os.path.join(dir_path, import_file)
        all_data = pd.read_excel(import_file)

        # Get all the masses.
        masses = {}
        for _, row in all_data.iterrows():
            key = snake_case(row['assay'])
            masses[key] = row['test_mass']

        # Aggregate all data for a sample
        data = all_data.groupby('sample_id', as_index=False).first()

        # Fill in sample details.
        for _, row in data.iterrows():

            # Get sample data as a dictionary.
            sample_data = {**row.to_dict(), **limits}
            sample_id = sample_data['sample_id']
            if not sample_id: # FIXME: Skip nan
                continue

            # Calculate terpene and cannabinoid results.
            sample_data = calculate_results(
                sample_data,
                analysis='terpenes',
                mass=masses['terpenes'],
                dilution_factor=40
            )
            sample_data = calculate_results(
                sample_data,
                analysis='cannabinoids',
                mass=masses['potency'],
                dilution_factor=40*50
            )

            # Iterate over worksheet pages to fill-in
            # all occurrences of each reference 1 by 1.
            template_workbook = openpyxl.load_workbook(coa_template_copy, keep_vba=True)
            fill_jinja_references(template_workbook, sample_data)

            # FIXME: Get output pages dynamically
            try:
                ws_index_list = get_worksheet_indexes(template_workbook, output_pages)
            except:
                if len(output_pages) == 3:
                    ws_index_list = [3, 4, 5]
                else:
                    ws_index_list = [3, 4, 5, 6]

            # Save the rendered template, temporarily.
            abs_render_file = os.path.join(dir_path, f'CoAs/{sample_id}.xlsm')
            template_workbook.save(abs_render_file)

            # Future: Insert signatures
            # Future: Insert QR Code
            # Future: Touch up the CoA.
                # ws.oddHeader.left.text = "Page &[Page] of &N"
                # Mark failures as red
                # a1.font = Font(color="FF0000", italic=True) # the change only affects A1

            # Create a PDF.
            output_file = f'CoAs/{sample_id}.pdf'
            excel = create_coa_pdfs(abs_render_file, ws_index_list, output_file)

            # Future: Upload the PDF.        
            # Future: Create download link and short link for the CoA.
            # Future: Upload the CoA data.
                # - See https://cloud.google.com/functions/docs/writing/http#uploading_files_via_cloud_storage

    # Remove temporary files.
    # os.remove(abs_render_file)
    os.remove(coa_template_copy)

    # Ensure Excel is visible.
    excel.ScreenUpdating = True
    excel.DisplayAlerts = True
    excel.EnableEvents = True


# TODO:
# def insert_qr_code(sheet, coords, url):
#     """Insert a QR code into a CoA template.
#     Args:
#         sheet (Worksheet): The worksheet to insert the QR Code
#         ref (str): The location to insert the QR code.
#         url (str): The URL that the QR code should link.
#     Returns:
#     """
#     qr = qrcode.QRCode(
#         version=1,
#         error_correction=qrcode.constants.ERROR_CORRECT_L,
#         box_size=10,
#         border=4,
#     )
#     qr.add_data(url)
#     qr.make(fit=True)
#     img = qr.make_image(fill_color='black', back_color='white')
#     img.save('qr_code.png')
#     logo = Image('qr_code.png')
#     logo.height = 150
#     logo.width = 150
#     sheet.add_image(logo, coords)
#     # workbook.save(filename="hello_world_logo.xlsx")


def read_worksheet(workbook, sheetname='Upload'):
    """Read the imported data, iterating over the rows and
    getting value from each cell in row.
    Args:
        path (str or InMemoryFile): An Excel workbook to read.
        filename (str): The name of the worksheet to upload.
    Returns:
    """
    sheet = workbook[sheetname]
    headers = get_worksheet_headers(sheet)
    return get_worksheet_data(sheet, headers)


def run_generate_coas():
    """Call `generate_coas` from an Excel workbook with xlwings."""

    # Initialize the workbook
    book = xlwings.Book.caller()
    worksheet = book.sheets.active
    config_sheet = book.sheets['cannlytics.conf']
    config = get_worksheet_data_block(config_sheet, 'A1', expand='table')
    show_status_message(
        worksheet,
        coords=config['status_cell'],
        message='Generating CoAs...',
        background=config['success_color'],
    )

    # Get the parameters.
    import_file = worksheet.range(config['import_file_cell']).value
    if import_file is None:
        show_status_message(
            worksheet,
            coords=config['status_cell'],
            message='Please provide an import file.',
            background=config['error_color']
        )
        return
    coa_template = config['coa_template']
    render_file = config['render_file']
    output_pages = worksheet.range(config['output_pages_cell']).value
    output_pages = output_pages.split(',')
    output_pages = [x.strip() for x in output_pages]
    limits = get_analyte_limits()

    # Generate the CoAs.
    generate_coas(
        import_file,
        output_pages=output_pages,
        coa_template=coa_template,
        # render_file=render_file,
        limits=limits
    )
    show_status_message(
        worksheet,
        coords=config['status_cell'],
        message='Generated CoAs',
        background=config['success_color'],
    )


def show_status_message(sheet, coords, message, background=None, color=None):
    """Show a status message in an Excel spreadsheet.
    Args:
        sheet (Sheet): The sheet where the status message will be written.
        coords (str): The location of the status message.
        message (str): A status message to write to Excel.
        background (tuple): Optional background color.
        color (tuple): Optional font color.
    """
    sheet.range(coords).value = message
    if background:
        sheet.range(coords).color = literal_eval(background)
    if color:
        sheet.range(coords).api.Font.Color = rgb_to_int(literal_eval(color))


def snake_case(string):
    """Turn a given string to snake case.
    Handles CamelCase, replaces known special characters with
    preferred namespaces, replaces spaces with underscores,
    and removes all other nuisance characters.
    Args:
        string (str): The string to turn to snake case.
    Returns"
        (str): A snake case string.
    """
    key = string.replace(' ', '_')
    key = key.replace('&', 'and')
    key = key.replace('%', 'percent')
    key = key.replace('#', 'number')
    key = key.replace('$', 'dollars')
    key = key.replace('/', '_')
    key = key.replace(r'\\', '_')
    key = sub('[!@#$%^&*()[]{};:,./<>?\|`~-=+]', ' ', key)
    keys = findall(r'[A-Z]?[a-z]+|[A-Z]{2,}(?=[A-Z][a-z]|\d|\W|$)|\d+', key)
    return '_'.join(map(str.lower, keys))


if __name__ == '__main__':   

    # Declare command line arguments.
    parser = argparse.ArgumentParser(description='CoA Generation')
    parser.add_argument('--files', action='store', dest='files', default=['export.xlsx'])
    parser.add_argument('--pages', action='store', dest='pages', default=['Page 1', 'Page 2', 'Page 3', 'Page 4'])
    parser.add_argument('--template', action='store', dest='template', default='coa_template.xlsm')
    args = parser.parse_args()

    # Format files as a list.
    files = args.files.replace('[', '').replace(']', '').split(',')
    files = [x.strip() for x in files]

    # Format pages as a list.
    pages = args.files.replace('[', '').replace(']', '').split(',')
    pages = [x.strip() for x in pages]

    # Read in analyte limits.
    limits = get_analyte_limits()

    # TEST:
    # files = './Export Files'
    # pages = ['Page 1', 'Page 2', 'Page 3', 'Page 4']

    # Optionally specify a folder instead of a list of files.
    # If its a directory, get all .xlsx files as import files.
    try:
        # dir_path = os.path.dirname(os.path.realpath(__file__))
        # full_file = os.path.join(dir_path, files)
        potential_dir = files[0]
        if os.path.exists(os.path.dirname(potential_dir)):
            import_files = []
            for file in os.listdir(potential_dir):
                if file.endswith('.xlsx'):
                    filename = os.path.join(potential_dir, file)
                    import_files.append(filename)
            files = import_files
    except TypeError:
        pass


    # Optional: Enter image files.

    # Test CoA generation.
    generate_coas(
        files,
        output_pages=pages,
        # coa_template=args.template,
        limits=limits
    )
