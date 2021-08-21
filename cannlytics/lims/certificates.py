"""
Certificate of Analysis (CoA) Generation | Cannlytics

Author: Keegan Skeate <keegan@cannlytics.com>  
Created: 2/6/2021  
Updated: 8/9/2021  
License: MIT LIcense <https://opensource.org/licenses/MIT>

TODO:
    - Allow users to create CoA's with a myriad of templates! (.docx, .xlsx, etc.)
    - Import any docx template styed with jinga-style formatting and let'r'rip.
    - Use your own templates or download the started templates and customize them to your heart's galore.
"""
# Standard imports
import os
import environ
from pathlib import Path
from shutil import copyfile

try:

    # External packages
    import openpyxl
    from openpyxl.drawing.image import Image
    import pandas as pd
    # import qrcode

    # Internal imports
    from cannlytics.firebase import (
        get_document,
        update_document,
        download_file,
    )
    from cannlytics.utils.utils import snake_case

except:
    pass

def create_coa():
    """
    Creates a certificate of analysis.
    """

    # TODO: Get sample details.

    # Get project details.

    # Get client details.

    # Get template.

    # Fill template
        # - If all of the analyses in a table have the same units,
        # then put the units in a footnote beneath. Otherwise,
        # append the units to the result.

        # Wishlist: Add initials of analyst.
        # Wishlist: Add date and time of analysis.
        # Wishlist: Add SoP, method, instrument.
        # Wishlist: Option to add dilution factor to CoA.

    # Create PDF in /tmp

    # Upload PDF to storage

    # Create short link (preferably create before upload?)

    # Create QR code

    # Update PDF, re-upload.

    return NotImplementedError


def delete_coa():
    """
    Delete a certificate of analysis from storage.
    """
    # Remove certificate from storage.

    # delete_file(bucket_name, blob_name)

    # Remove the certificate data.

    # update_document(ref, {'coa_url': '', 'coa_ref': ''})

    return NotImplementedError


def email_coas():
    """
    Email certificates of analysis to their recipients.
    """
    # Get the certificate data.

    # Email links (optional attachments) to the contacts.

    return NotImplementedError


def text_coas():
    """
    Text certificates of analysis to their recipients.
    """
    # Get the certificate data.

    # Text links to the contacts.

    return NotImplementedError


def send_coas():
    """
    Email and/or text certificates of analysis to their recipients.
    """
    # Email and / or text CoAs using email_coas and text_coas.

    return NotImplementedError


def get_coa_urls():
    """
    Get CoA URLs for given samples or projects.
    """
    # Get certificate data and return the short links.

    return NotImplementedError


def merge_coas():
    """
    Merge two CoAs together, either existing templates
    or existing PDFs.
    """
    # Get certificate data for the desired CoAs.

    # Download the existing PDFs for the certificates.

    # Merge the downloaded PDFs.

    # Upload the merged PDF to storage.

    # Create short link.

    # Create QR code.

    # Re-upload the PDF.

    return NotImplementedError



def approve_coa():
    """
    Creates a certificate of analysis.
    """
    # Verify the user's pin.

    # Get the user's signature.

    # Insert the user's signature on the certificate.

    print('Approving CoA...')
    return NotImplementedError


def review_coa(env_file='.env', signature_dest='./tmp/signature.png'):
    """
    Creates a certificate of analysis.
    """
    # TODO: Validate the user's pin to get their UID.
    uid = ''

    # Get the user's signature (if not already downloaded?).
    env = environ.Env()
    env.read_env(env_file)
    bucket_name = env('FIREBASE_STORAGE_BUCKET')
    signature_data = get_document(f'users/{uid}/user_settings/signature')
    download_file(bucket_name, signature_data['signature_ref'], signature_dest)

    # Insert the signature into the CoA template.

    # Create the PDF.

    # Upload the PDF to storage.

    # Save the reviewer data.

    print('Reviewing CoA...')
    return NotImplementedError


def upload_coa():
    """
    Upload a certificate of analysis to storage.
    """
    # Read file data.

    # Upload file data to storage.

    # Update certificate data.

    return NotImplementedError


def upload_coa_template():
    """
    Upload a certificate of analysis template to storage.
    """
    # Read in the file data.

    # Upload the template to storage.

    # Create a download link for the template.

    # Save the template data.

    return NotImplementedError


def create_short_url():
    """
    Create a short URL for a CoA link.
    """
    return NotImplementedError


# def insert_qr_code(sheet, coords, url):
#     """Insert a QR code into a CoA template.
#     Args:
#         sheet (Worksheet): The worksheet to insert the QR Code
#         ref (str): The location to insert the QR code.
#         url (str): The URL that the QR code should link.
#     Returns:
#     """
#     qr = qrcode.QRCode(
#         version=1,
#         error_correction=qrcode.constants.ERROR_CORRECT_L,
#         box_size=10,
#         border=4,
#     )
#     qr.add_data(url)
#     qr.make(fit=True)
#     img = qr.make_image(fill_color='black', back_color='white')
#     img.save('qr_code.png')
#     logo = Image('qr_code.png')
#     logo.height = 150
#     logo.width = 150
#     sheet.add_image(logo, coords)
#     # workbook.save(filename="hello_world_logo.xlsx")

#-----------------------------------------------------------------------
# FIXME: Generate CoA routine
#-----------------------------------------------------------------------

def generate_coas(
        import_files,
        output_pages,
        coa_template='./coa_template.xlsm',
        # render_file='./CoAs/coa_render.xlsm',
        limits={}
):
    """Generate certificates of analysis.
    Args:
        import_files (list): A list of files to import.
        output_pages (list): A list of pages to include in the PDF.
        coa_template (str): The path of the CoA Template.
        limits (dict): A dictionary of limits and LOQ for analytes.
    """

    # Create CoA folder if one does not exist.
    Path('CoAs').mkdir(parents=True, exist_ok=True)
    dir_path = os.path.dirname(os.path.realpath(__file__))

    # Create a copy of the template.
    abs_coa_template = os.path.join(dir_path, coa_template)
    coa_template_copy = abs_coa_template.replace('.xlsm', '_copy.xlsm')
    copyfile(abs_coa_template, coa_template_copy)

    # Iterate over all import files.
    for import_file in import_files:

        # Get all sample, results, client, etc. data.
        # abs_import_file = os.path.join(dir_path, import_file)
        all_data = pd.read_excel(import_file)

        # Get all the masses.
        masses = {}
        for _, row in all_data.iterrows():
            key = snake_case(row['assay'])
            masses[key] = row['test_mass']

        # Aggregate all data for a sample
        data = all_data.groupby('sample_id', as_index=False).first()

        # Fill in sample details.
        for _, row in data.iterrows():

            # Get sample data as a dictionary.
            sample_data = {**row.to_dict(), **limits}
            sample_id = sample_data['sample_id']
            if not sample_id: # FIXME: Skip nan
                continue

            # Calculate terpene and cannabinoid results.
            sample_data = calculate_results(
                sample_data,
                analysis='terpenes',
                mass=masses['terpenes'],
                dilution_factor=40
            )
            sample_data = calculate_results(
                sample_data,
                analysis='cannabinoids',
                mass=masses['potency'],
                dilution_factor=40*50
            )

            # Iterate over worksheet pages to fill-in
            # all occurrences of each reference 1 by 1.
            template_workbook = openpyxl.load_workbook(coa_template_copy, keep_vba=True)
            fill_jinja_references(template_workbook, sample_data)

            # FIXME: Get output pages dynamically
            try:
                ws_index_list = get_worksheet_indexes(template_workbook, output_pages)
            except:
                if len(output_pages) == 3:
                    ws_index_list = [3, 4, 5]
                else:
                    ws_index_list = [3, 4, 5, 6]

            # Save the rendered template, temporarily.
            abs_render_file = os.path.join(dir_path, f'CoAs/{sample_id}.xlsm')
            template_workbook.save(abs_render_file)

            # Future: Insert signatures
            # Future: Insert QR Code
            # Future: Touch up the CoA.
                # ws.oddHeader.left.text = "Page &[Page] of &N"
                # Mark failures as red
                # a1.font = Font(color="FF0000", italic=True) # the change only affects A1

            # Create a PDF.
            output_file = f'CoAs/{sample_id}.pdf'
            # FIXME:
            excel = create_coa_pdfs(abs_render_file, ws_index_list, output_file)

            # Future: Upload the PDF.        
            # Future: Create download link and short link for the CoA.
            # Future: Upload the CoA data.
                # - See https://cloud.google.com/functions/docs/writing/http#uploading_files_via_cloud_storage

    # Remove temporary files.
    # os.remove(abs_render_file)
    os.remove(coa_template_copy)

    # Ensure Excel is visible.
    excel.ScreenUpdating = True
    excel.DisplayAlerts = True
    excel.EnableEvents = True


# FIXME: Do it without Excel :(
def create_coa_pdfs(render_file, ws_index_list, output_file, tight=False):
    """Generate PDFs for rendred CoAs.
    Args:
        render_file (str): The path of the rendred workbook.
        ws_index_list (list): A list of the worksheet indexes to include in the PDF.
        output_file (str): The name of the output PDF (expected .pdf extension).
        tight (bool): Optional, default False, choice to scale all pages.
    Returns:
        (COM Object): A Microsoft Excel Client to alter later.
    """
    client = win32com.client.Dispatch('Excel.Application')
    client.Visible = False
    client.ScreenUpdating = False
    client.DisplayAlerts = False
    client.EnableEvents = False
    dir_path = os.path.dirname(os.path.realpath(__file__))
    input_file = os.path.join(dir_path, render_file)
    pdf_file = os.path.join(dir_path, output_file)
    workbook = client.Workbooks.Open(input_file)
    if tight:
        print_area = 'A1:G50'
        for index in ws_index_list:
            worksheet = workbook.Worksheets[index]
            worksheet.PageSetup.Zoom = False
            worksheet.PageSetup.FitToPagesTall = 1
            worksheet.PageSetup.FitToPagesWide = 1
            worksheet.PageSetup.PrintArea = print_area
    try:
        workbook.WorkSheets(ws_index_list).Select()
    except:
        workbook.Worksheets(ws_index_list).Select()
    workbook.ActiveSheet.ExportAsFixedFormat(0, pdf_file)
    workbook.Close(False)
    return client

#-----------------------------------------------------------------------
# Internal functions
#-----------------------------------------------------------------------

def fill_jinja_references(workbook, data):
    """Fill-in Jinga-style references, iterating over all
    worksheet pages to fill-in all occurrences of each reference 1 by 1.
    Args:
        workbook (Workbook): A Workbook object.
        data (dict): A dictionary of context data.
    """
    context = { snake_case(key): values for key, values in data.items() }
    sheets = workbook.worksheets
    for sheet in sheets:
        refs = get_jinja_references(sheet)
        for key, cells in refs.items():
            clean_key = snake_case(key)
            value = context.get(clean_key, 0)
            for cell in cells:
                if value:
                    sheet[cell] = str(value)
                else:
                    sheet[cell] = 'MISSING'


def get_jinja_references(sheet):
    """Get Jinja-style references in an Excel worksheet.
    Args:
        sheet (Worksheet): A worksheet to get references from.
    Returns:
        (dict): A dictionary of variables to find references.
    """
    refs = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            try:
                if value.startswith('{{'):
                    ref = cell.column_letter + str(cell.row)
                    variable = value.replace('{{', '').replace('}}', '').strip()
                    existing_refs = refs.get(variable, [])
                    refs[variable] = [*existing_refs, *[ref]]
                elif value.startswith('{% for'): # Optional: Handle tables
                    return NotImplementedError
            except AttributeError:
                continue
    return refs


def get_worksheet_data(sheet, headers):
    """Get the data of a worksheet.
    Args:
        sheet (Worksheet): An openpyx; Excel file object.
        headres (list): A list of headers to map the values.
    Returns:
        list(dict): A list of dictionaries.
    """
    data = []
    for row in sheet.iter_rows():
        values = {}
        for key, cell in zip(headers, row):
            values[key] = cell.value
        data.append(values)
    return data


def get_worksheet_data_block(sheet, coords, expand=None):
    """Get a data block.
    Args:
        sheet (Sheet): The worksheet containing the data block.
        coords (str): The inclusive coordinates of the data block.
        expand (str): Optionally expand the range of values.
    Returns
        (dict): A dictionary of the data in the data block.
    """
    data = {}
    values = sheet.range(coords).options(expand=expand).value
    for item in values:
        key = snake_case(item[0])
        value = item[1]
        data[key] = value
    return data


def get_worksheet_headers(sheet):
    """Get the headres of a worksheet.
    Args:
        sheet (Worksheet): An openpyx; Excel file object.
    Returns:
        headers (list): A list of header strings.
    """
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    return headers


def get_worksheet_indexes(wb, output_pages):
    """Get the indexes for a list of sheet names in a given workbook.
    Args:
        wb (Workbook): The workbook at hand.
        pages (list): A list of pages to find indexes.
    """
    ws_index_list = []
    for page in output_pages:
        index = wb.worksheets.index(wb[page])
        ws_index_list.append(index + 1)
    return ws_index_list


if __name__ == '__main__':

    import argparse

    # Declare command line arguments.
    parser = argparse.ArgumentParser(description='CoA Generation')
    parser.add_argument('--files', action='store', dest='files', default=['export.xlsx'])
    parser.add_argument('--pages', action='store', dest='pages', default=['Page 1', 'Page 2', 'Page 3', 'Page 4'])
    parser.add_argument('--template', action='store', dest='template', default='coa_template.xlsm')
    args = parser.parse_args()

    # Format files as a list.
    files = args.files.replace('[', '').replace(']', '').split(',')
    files = [x.strip() for x in files]

    # Format pages as a list.
    pages = args.files.replace('[', '').replace(']', '').split(',')
    pages = [x.strip() for x in pages]

    # Read in analyte limits.
    limits = get_analyte_limits()

    # TEST:
    # files = './Export Files'
    # pages = ['Page 1', 'Page 2', 'Page 3', 'Page 4']

    # Optionally specify a folder instead of a list of files.
    # If its a directory, get all .xlsx files as import files.
    try:
        # dir_path = os.path.dirname(os.path.realpath(__file__))
        # full_file = os.path.join(dir_path, files)
        potential_dir = files[0]
        if os.path.exists(os.path.dirname(potential_dir)):
            import_files = []
            for file in os.listdir(potential_dir):
                if file.endswith('.xlsx'):
                    filename = os.path.join(potential_dir, file)
                    import_files.append(filename)
            files = import_files
    except TypeError:
        pass


    # Optional: Enter image files.

    # Test CoA generation.
    generate_coas(
        files,
        output_pages=pages,
        # coa_template=args.template,
        limits=limits
    )