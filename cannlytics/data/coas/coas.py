"""
CoADoc | A Certificate of Analysis (COA) Parser
Copyright (c) 2022-2023 Cannlytics

Authors:
    Keegan Skeate <https://github.com/keeganskeate>
    Candace O'Sullivan-Sutherland <https://github.com/candy-o>
Created: 7/15/2022
Updated: 9/16/2023
License: <https://github.com/cannlytics/cannlytics/blob/main/LICENSE>

Description:

    Certificates of analysis (CoAs) are abundant for cannabis cultivators,
    processors, retailers, and consumers too, but the data is often
    locked away. Rich, valuable laboratory data so close, yet so far
    away! CoADoc puts these vital data points in your hands by
    parsing PDFs and URLs, finding all the data, standardizing the data,
    and cleanly returning the data to you.

Future work:

    - [ ] Integrate `create_hash` into `save` and `standardize`.
    - [ ] Improve the `standardize` method.

"""
# Standard imports.
import ast
import base64
from datetime import datetime
import importlib
from io import BytesIO
import json
import operator
import os
import tempfile
from typing import Any, List, Optional

# External imports.
import cv2
import numpy as np
# try:
#     import openai
# except ImportError:
#     print('Unable to find `openai` package. This tool is used for parsing with AI.')
import openpyxl
import pandas as pd
import requests
import pdfplumber
from PIL import Image
from pypdf import PdfMerger
try:
    from pyzbar import pyzbar
    # from pyzbar.pyzbar import decode
except:
    print('Unable to import `zbar` library. This tool is used for decoding QR codes.')
try:
    from pytesseract import image_to_pdf_or_hocr
except ImportError:
    print('Unable to import `Tesseract` library. This tool is used for OCR.')
try:
    from wand.image import Image as magick_wand
    from wand.color import Color
except ImportError:
    print('Unable to import `ImageMagick` library. This tool is used for OCR.')

# Internal imports.
from cannlytics.data.data import create_hash, write_to_worksheet
from cannlytics.data.web import download_google_drive_file
from cannlytics.utils import (
    convert_to_numeric,
    dump_column,
    get_directory_files,
    sandwich_list,
    reorder_columns,
    snake_case,
    unzip_files,
)
from cannlytics.utils.constants import (
    ANALYSES,
    ANALYTES,
    CODINGS,
    DEFAULT_HEADERS,
    STANDARD_FIELDS,
)

# Lab and LIMS CoA parsing algorithms.
from cannlytics.data.coas.coa_ai import parse_coa_with_ai
from cannlytics.data.coas.algorithms.acs import ACS_LABS
from cannlytics.data.coas.algorithms.anresco import ANRESCO
from cannlytics.data.coas.algorithms.cannalysis import CANNALYSIS
from cannlytics.data.coas.algorithms.confidence import CONFIDENCE
from cannlytics.data.coas.algorithms.confidentcannabis import CONFIDENT_CANNABIS
from cannlytics.data.coas.algorithms.greenleaflab import GREEN_LEAF_LAB
from cannlytics.data.coas.algorithms.kaycha import KAYCHA_LABS
from cannlytics.data.coas.algorithms.mcrlabs import MCR_LABS
from cannlytics.data.coas.algorithms.sclabs import SC_LABS
from cannlytics.data.coas.algorithms.sonoma import SONOMA
from cannlytics.data.coas.algorithms.tagleaf import TAGLEAF
from cannlytics.data.coas.algorithms.terplife import TERPLIFE_LABS
from cannlytics.data.coas.algorithms.steephill import STEEPHILL
from cannlytics.data.coas.algorithms.veda import VEDA_SCIENTIFIC

# Labs and LIMS that CoADoc can parse.
LIMS = {
    'ACS Labs': ACS_LABS,
    'Anresco Laboratories': ANRESCO,
    'Cannalysis': CANNALYSIS,
    'Confidence Analytics': CONFIDENCE,
    'Confident Cannabis': CONFIDENT_CANNABIS,
    'Green Leaf Lab': GREEN_LEAF_LAB,
    'Kaycha Labs': KAYCHA_LABS,
    'MCR Labs': MCR_LABS,
    'SC Labs': SC_LABS,
    'Sonoma Lab Works': SONOMA,
    'TagLeaf LIMS': TAGLEAF,
    'TerpLife Labs': TERPLIFE_LABS,
    'Steep Hill': STEEPHILL,
    'Veda Scientific': VEDA_SCIENTIFIC,
}

# Default preferred order for DataFrame columns.
# TODO: Add `sample_hash` and `results_hash` to the beginning.
DEFAULT_COLUMN_ORDER = ['product_name', 'producer',
    'product_type', 'date_tested']

# Default nuisance columns to remove during standardization.
DEFAULT_NUISANCE_COLUMNS = ['received_by', 'sampled_by', 'Unnamed: 1',
    'None_method', 'loss_on_drying_moisture', '_3_5_grams',
    'index', 'other_analyses', 'wildcard']

# Default columns to apply codings and be treated as numeric.
DEFAULT_NUMERIC_COLUMNS = ['value', 'mg_g', 'lod', 'loq', 'limit', 'margin_of_error']

# Encountered Metrc prefixes used to identify Metrc IDs.
METRC_PREFIXES = ['1A40']

# A custom `ValueError` message to raise when no known LIMS is identified.
UNIDENTIFIED_LIMS = 'COA not recognized as a COA from a known lab: '
UNIDENTIFIED_LIMS += ', '.join([f'"{x}"' for x in LIMS.keys()])


def convert_pdf_to_images(
        filename: str,
        output_path: str,
        resolution: Optional[int] = 300,
    ) -> List[str]:
    """Convert a PDF into images, with images named as `{filename}-{#}.png`.
    The function replaces the images' alpha channels with a white background.
    Args:
        filename (str): The name of a PDF to convert to images.
        output_path (str): A directory where the images should be generated.
        resolution (int): The resolution of the generated images, 300 by
            default (optional.)
    Returns:
        (list): Returns a list of the filenames of the images generated.
    Authors: Thibaut Mattio, Averner <https://stackoverflow.com/a/42525093/5021266>
    License: CC BY-SA 4.0 <https://creativecommons.org/licenses/by-sa/4.0/>
    """
    image_files = []
    os.environ['MAGICK_TMPDIR'] = output_path
    all_pages = magick_wand(filename=filename, resolution=resolution)
    for i, page in enumerate(all_pages.sequence):
        with magick_wand(page) as img:
            img.format = 'png'
            img.background_color = Color('white')
            img.alpha_channel = 'remove'
            image_filename = os.path.basename(filename)
            image_filename = os.path.splitext(image_filename)[0]
            image_filename = '{}-{}.png'.format(image_filename, i)
            image_filename = os.path.join(output_path, image_filename)
            img.save(filename=image_filename)
            image_files.append(image_filename)
    # Help `ImageMagick` remove temporary files.
    for i in os.listdir(output_path):
        magick_path = os.path.join(output_path, i)
        if os.path.isfile(magick_path) and i.startswith('magick-'):
            os.remove(magick_path)
    return image_files


class CoADoc:
    """Parse data from certificate of analysis (CoA) PDFs or URLs."""

    def __init__(
            self,
            lims: Optional[Any] = None,
            analyses: Optional[dict] = None,
            analytes: Optional[dict] = None,
            codings: Optional[dict] = None,
            column_order: Optional[list] = None,
            nuisance_columns: Optional[list] = None,
            numeric_columns: Optional[list] = None,
            standard_fields: Optional[dict] = None,
            google_maps_api_key: Optional[str] = None,
            headers: Optional[dict] = None,
            init_all: Optional[bool] = True,
        ) -> None:
        """Initialize CoA parser.
        Args:
            analyses (dict): A dictionary of analyses,
                standard analyses are used by default.
            analytes (dict): A dictionary of analytes,
                standard analytes are used by default.
            codings (dict): A dictionary of value codings,
                standard codings are used by default.
            standard_fields (dict): A dictionary of field keys,
                standard fields are used by default.
            headers (dict): Headers for HTTP requests,
                standard headers are used by default.
            lims (str or dict): Specific LIMS to parse CoAs,
                all verified LIMS are checked by default,
                until a matching LIMS is found.
            init_all (bool): Initialize all of the parsing routines
                for all verified LIMS.
            google_maps_api_key (str): A Google Maps API key
                if you also want GIS data included.
        """
        # Setup.
        self.driver = None
        self.options = None
        self.session = None
        self.service = None

        # Define analyses.
        self.analyses = analyses
        if analyses is None:
            self.analyses = ANALYSES

        # Define analytes.
        self.analytes = analytes
        if analytes is None:
            self.analytes = ANALYTES

        # Define codings.
        self.codings = codings
        if codings is None:
            self.codings = CODINGS

        # Define default column / field order.
        self.column_order = column_order
        if column_order is None:
            self.column_order = DEFAULT_COLUMN_ORDER

        # Define default nuisance columns / fields to remove.
        self.nuisance_columns = nuisance_columns
        if nuisance_columns is None:
            self.nuisance_columns = DEFAULT_NUISANCE_COLUMNS

        # Define default columns / fields to treat as numeric.
        self.numeric_columns = numeric_columns
        if numeric_columns is None:
            self.numeric_columns = DEFAULT_NUMERIC_COLUMNS 

        # Define fields.
        self.fields = standard_fields
        if standard_fields is None:
            self.fields = STANDARD_FIELDS

        # Define headers.
        self.headers = headers
        if headers is None:
            self.headers = DEFAULT_HEADERS

        # Define LIMS.
        self.lims = lims
        if lims is None:
            self.lims = LIMS

        # Optional Google Maps integration to retrieve GIS data.
        self.google_maps_api_key = google_maps_api_key
    
        # Assign all of the parsing routines.
        if init_all:
            for values in self.lims.values():
                script = values['coa_algorithm'].replace('.py', '')
                module = f'cannlytics.data.coas.algorithms.{script}'
                entry_point = values['coa_algorithm_entry_point']
                setattr(self, entry_point, importlib.import_module(module))

    def aggregate(
            self,
            datafiles,
            output: Optional[str] = None,
            sheet_name: Optional[str] = None,
            **kwargs,
        ) -> pd.DataFrame:
        """Aggregate multiple datafiles into a single dataset.
        Args:
            datafiles (iterable): A list of filenames, DataFrames, or 
                lists of dictionaries to aggregate into a single DataFrame.
            output (str): A path or file where data should be output (optional).
            sheet_name (str): A specific sheet to read data (optional).
            kwargs (Keywords): Keywords to pass to `save`.
        Returns:
            (list): Returns a list of data dictionaries.
        """
        # Aggregate DataFrames.
        agg = pd.DataFrame()
        for d in datafiles:
            if isinstance(d, str):
                df = pd.read_excel(d, sheet_name=sheet_name)
                agg = pd.concat([agg, df])
            elif isinstance(d, pd.DataFrame):
                agg = pd.concat([agg, d])
            elif isinstance(d, list):
                agg = pd.concat([agg, pd.DataFrame(d)])

        # Replace NaN values with `None`.
        agg = agg.where(pd.notnull(agg), None)

        # Remove duplicates.
        agg.drop_duplicates('sample_hash', inplace=True)

        # Create aggregation columns.
        agg['results'] = agg['results'].apply(dump_column)
        agg = agg.loc[agg['results'].notnull()]

        # TODO: Also handle `images` and `coa_urls`.
        # agg['images'] = agg['images'].apply(dump_column)
        # agg['coa_urls'] = agg['coa_urls'].apply(dump_column)

        # Re-hash the dataset.
        datafile_hash = create_hash(agg)

        # Save the data.
        datafile = None
        if os.path.isdir(output):
            datafile = f'{output}/{datafile_hash}.xlsx'
        elif os.path.isfile(output):
            datafile = output
        elif output is True:
            datafile = f'{datafile_hash}.xlsx'
        if datafile:
            self.save(agg, datafile, **kwargs)
        return agg

    def decode_pdf_qr_code(
            self,
            page: Any,
            img: Any,
            resolution: Optional[int] = 300,
        ) -> list:
        """Decode a PDF QR Code from a given image.
        Args:
            page (Page): A pdfplumber Page containing the image.
            img (Image): A pdfplumber Image.
            resolution (int): The resolution to render the QR code,
                `300` by default (optional).
        Returns:
            (list): The QR code data.
        """
        y = page.height
        bbox = (img['x0'], y - img['y1'], img['x1'], y - img['y0'])
        crop = page.crop(bbox)
        obj = crop.to_image(resolution=resolution)
        return pyzbar.decode(obj.original)

    def find_pdf_qr_code_url(
            self,
            pdf: Any,
            image_index: Optional[int] = None,
            page_index: Optional[int] = 0,
            resolution: Optional[int] = 300,
        ) -> str:
        """Find the QR code given a CoA PDF or page.
        If no `image_index` is provided, then all images are tried to be
        decoded until a QR code is found. If no QR code is found, then a
        `IndexError` is raised.
        Args:
            pdf (PDF or Page): A pdfplumber PDF or Page.
            image_index (int): A known image index for the QR code.
            page_index (int): The page to search, 0 by default (optional).
            resolution (int): The resolution to render the QR code,
                `300` by default (optional).
        Returns:
            (str): The QR code URL.
        """
        image_data = None
        if isinstance(pdf, str):
            pdf_file = pdfplumber.open(pdf)
            page = pdf_file.pages[page_index]
        elif isinstance(pdf, pdfplumber.pdf.PDF):
            page = pdf.pages[page_index]
        else:
            page = pdf
        if image_index:
            img = page.images[image_index]
            decoded_image = self.decode_pdf_qr_code(page, img, resolution)
            image_data = decoded_image[0].data.decode('utf-8')
        else:
            image_range = sandwich_list(page.images)
            for img in image_range:
                try:
                    decoded_image = self.decode_pdf_qr_code(page, img, resolution)
                    image_data = decoded_image[0].data.decode('utf-8')
                    if image_data:
                        break
                except:
                    continue
        return image_data

    def find_metrc_ids(
            self,
            doc: Any,
            prefixes: Optional[List[str]] = None,
            id_length: Optional[int] = 24,
        ) -> List[str]:
        """Find any Metrc IDs that may be in a given CoA PDF by searching
        for known Metrc prefixes and assuming that the ID is 24 characters.
        Arg:
            doc (str or PDF or Page): A filename, PDF, or Page.
            prefixes (list): A list of Metrc prefixes (optional).
            id_length (int): The length of a Metrc ID, 24 by default (optional).
        Returns:
            (list): Returns a list of Metrc IDs.
        """
        if isinstance(doc, str):
            pdf_file = pdfplumber.open(doc)
            pages = pdf_file.pages
        elif isinstance(doc, pdfplumber.pdf.Page):
            pages = [doc]
        else:
            pages = doc.pages
        if prefixes is None:
            prefixes = METRC_PREFIXES
        metrc_ids = []
        for prefix in prefixes:
            for page in pages:
                text = page.extract_text().replace('\n', '')
                words = text.split(' ')
                ids = [x for x in words if x.startswith(prefix) and \
                    len(x) == id_length]
                metrc_ids.extend(ids)
        return list(set(metrc_ids))

    def get_metrc_results(self, metrc_id: str) -> dict:
        """Get Metrc lab results that have been archived
        in the Cannlytics library via the Cannlytics API.
        Args:
            metrc_id (str): A Metrc UID.
        Returns:
            (dict): The sample data.
        """
        url = f'https://cannlytics.com/api/data/results?metrc_id={metrc_id}'
        response = requests.get(url)
        try:
            return response.json()['data']
        except (json.decoder.JSONDecodeError, KeyError):
            return None

    def get_page_rows(self, page: Any, **kwargs) -> list:
        """Get the rows a given page.
        Args:
            page (Page): A pdfplumber page containing rows to extract.
        Returns:
            (list): A list of text.
        """
        txt = page.extract_text(**kwargs)
        txt = txt.replace('\xa0\xa0', '\n').replace('\xa0', ',')
        return txt.split('\n')

    def get_pdf_creation_date(self, pdf: Any) -> str:
        """Get the creation date of a PDF in ISO format.
        Args:
            pdf (PDF): A pdfplumber PDF.
        Returns:
            (str): An ISO formatted date.
        """
        if isinstance(pdf, str):
            pdf_file = pdfplumber.open(pdf)
        else:
            pdf_file = pdf
        try:
            date = pdf_file.metadata['CreationDate'].split('D:')[-1]
        except KeyError:
            return None
        isoformat = f'{date[0:4]}-{date[4:6]}-{date[6:8]}'
        isoformat += f'T{date[8:10]}:{date[10:12]}:{date[12:14]}'
        return isoformat

    def get_pdf_image_data(
            self,
            page: Any,
            bbox: Optional[tuple] = None,
            image_index: Optional[int] = 0, 
            resolution: Optional[int] = 300,
        ) -> str:
        """Get the image data for a given PDF page image.
        Args:
            page (Page): A pdfplumber Page.
            image_index (int): The index of the image, 0 by default (optional).
            resolution (int): The resolution for the image, 300 by default (optional).
        Returns:
            (str): The image data.
        """
        if bbox is None:
            y = page.height
            img = page.images[image_index]
            bbox = (img['x0'], y - img['y1'], img['x1'], y - img['y0'])
        crop = page.crop(bbox)
        obj = crop.to_image(resolution=resolution)
        buffered = BytesIO()
        obj.save(buffered, format='JPEG')
        img_str = base64.b64encode(buffered.getvalue())
        return img_str.decode('utf-8')

    def save_image_data(
            self,
            image_data: str,
            image_file: Optional[str] ='image.png',
        ):
        """Save image data to a file.
        Args:
            image_data (str): The image data.
            image_file (str): A filename for the image.
        """
        image_bytes = base64.b64decode(image_data)
        image_io = BytesIO(image_bytes)
        image = Image.open(image_io)
        image.save(image_file)

    def identify_lims(
            self,
            doc: Any,
            lims: Optional[Any] = None,
            temp_path: Optional[str] = '/tmp',
        ) -> str:
        """Identify if a CoA was created by a common LIMS.
        Search all of the text of the LIMS name or URL.
        If no LIMS is identified from the text, then the images
        are attempted to be decoded, searching for a QR code URL.
        Args:
            doc (str, PDF or Page): A URL or a pdfplumber PDF or Page.
            lims (str or dict): The name of a specific LIMS or a
                dictionary of known LIMS.
            temp_path (str): A temporary directory to store any online PDFs
                if needed for identification, `/tmp` by default (optional).
        Returns:
            (str): Returns LIMS name if found, otherwise returns `None`.
        """
        # Search all of the text of the LIMS name or URL.
        known = None
        if isinstance(doc, str):

            # Handle shortened URLs.
            text = doc
            if doc.startswith('https://tinyurl'):
                response = requests.get(doc, headers=DEFAULT_HEADERS)
                text = response.url
        
            # Handle Google Drive PDFs.
            if text.startswith('https://drive.google'):
                temp_pdf = os.path.join(temp_path, 'coa.pdf')
                if not os.path.exists(temp_path): os.makedirs(temp_path)
                download_google_drive_file(text, temp_pdf)
                text = temp_pdf

            # Handle PDFs.
            try:
                pdf_file = pdfplumber.open(text)
                text = pdf_file.pages[0].extract_text()
            except (FileNotFoundError, OSError):
                pass
        
        # Handle PDFs passed directly.
        else:
            if isinstance(doc, pdfplumber.pdf.PDF):
                page = doc.pages[0]
            else:
                page = doc
            text = page.extract_text()

        # Handle custom LIMS.
        if lims is None:
            lims = LIMS
        
        # Iterate over all known LIMS looking for URLs and LIMS names.
        if isinstance(lims, str):
            try:
                if self.lims[lims]['url'] in text:
                    known = lims
            except KeyError:
                if lims.lower() in text.lower():
                    known = lims
        else:
            for key, values in lims.items():
                url = values.get('url')
                if url and url in text:
                    known = key
                    break
                lab = values.get('lims')
                if lab and lab in text:
                    known = key
                    break

        # If the LIMS is still unknown, then try to identify the LIMS
        # with any a QR code, if the COA is a PDF.
        if not known:
            try:
                qr_code_url = self.find_pdf_qr_code_url(doc)
            except:
                qr_code_url = None
            if qr_code_url:
                for key, values in lims.items():
                    url = values.get('url')
                    if url and url in qr_code_url:
                        known = key
                        break
        
        # Optional: Can we detect LIMS from a photo for a COA or a label?

        # Return any known LIMS.
        return known

    def parse(
            self,
            data: Any,
            cleanup: Optional[bool] = True,
            headers: Optional[dict] = {},
            kind: Optional[str] = 'url',
            lims: Optional[Any] = None,
            max_delay: Optional[float] = 7,
            persist: Optional[bool] = True,
            resolution: Optional[int] = 300,
            temp_path: Optional[str] = '/tmp',
            use_cached: Optional[bool] = False,
            verbose: Optional[bool] = False,
        ) -> list:
        """Parse all CoAs given a directory, a list of files,
        or a list of URLs.
        Args:
            data (str or list): A directory (str) or a list
                of PDF file paths or a list of CoA URLs.
            cleanup (bool): Whether or not to remove the files generated
                during OCR, `True` by default (optional).
            headers (dict): Headers for HTTP requests (optional).
            kind (str): The kind of CoA input, PDF or URL, `url`
                by default (optional).
            lims (str or dict): Specific LIMS to parse CoAs (optional).
            max_delay (float): The maximum number of seconds to wait
                for the page to load (optional).
            persist (bool): Whether to persist the driver
                and / or session between CoA parses, the
                default is `True`, with any driver and session
                being closed at the end (optional).
            temp_path (str): A temporary directory to store files used
                during PDF OCR, `/tmp` by default (optional).
            resolution (int): The resolution of rendered PDF images,
                300 by default (optional).
        Returns:
            (list): Returns a list of all of the PDFs.
        """
        coas, docs = [], []

        # Parse a URL, PDF path, or .zip folder.
        if isinstance(data, str):

            # Parse a URL to an image of a COA or QR code.
            image_extensions = ['.png', '.jpg', '.jpeg']
            if 'https' in data and any(ext in data.lower() for ext in image_extensions):
                if verbose:
                    print('Parsing image URL.')
                image_url = self.scan(data)
                coa_data = self.parse_url(
                    image_url,
                    headers=headers,
                    lims=lims,
                    max_delay=max_delay,
                    persist=persist,
                    verbose=verbose,
                )
                coas.append(coa_data)

            # Parse a URL.
            elif 'https' in data:
                if verbose:
                    print('Parsing URL.')
                coa_data = self.parse_url(
                    data,
                    headers=headers,
                    lims=lims,
                    max_delay=max_delay,
                    persist=persist,
                    verbose=verbose,
                )
                coas.append(coa_data)

            # Parse a PDF.
            elif '.pdf' in data:
                if verbose:
                    print('Parsing PDF.')
                coa_data = self.parse_pdf(
                    data,
                    cleanup=cleanup,
                    headers=headers,
                    lims=lims,
                    max_delay=max_delay,
                    persist=persist,
                    resolution=resolution,
                    temp_path=temp_path,
                    use_cached=use_cached,
                    verbose=verbose,
                )
                coas.append(coa_data)

            # TODO: Handle paths to images of COAs and QR codes.

            # Parse a ZIP.
            elif '.zip' in data:
                if verbose:
                    print('Parsing ZIP.')
                doc_dir = unzip_files(data)
                docs = get_directory_files(doc_dir)

            # Parse a directory.
            else:
                if verbose:
                    print('Parsing directory.')
                docs = get_directory_files(data)

        # Handle a list of URLs, PDFs, and/or ZIPs.
        else:
            if verbose:
                print('Parsing list.')
            docs = data

        # Parse all of the PDFs.
        for doc in docs:

            # Parse a .zip folder.
            if '.zip' in doc and kind != 'url':
                if verbose:
                    print('Parsing ZIP.')
                doc_dir = unzip_files(doc)
                pdf_files = get_directory_files(doc_dir)
                for pdf_file in pdf_files:
                    coa_data = self.parse_pdf(
                        pdf_file,
                        cleanup=cleanup,
                        headers=headers,
                        lims=lims,
                        max_delay=max_delay,
                        persist=persist,
                        resolution=resolution,
                        temp_path=temp_path,
                        use_cached=use_cached,
                        verbose=verbose,
                    )
                    coas.append(coa_data)

            # Parse a URL.
            elif doc.startswith('http'):
                if verbose:
                    print('Parsing URL.')
                coa_data = self.parse_url(
                    doc,
                    headers=headers,
                    lims=lims,
                    max_delay=max_delay,
                    persist=persist,
                    verbose=verbose,
                )
            
            # Parse a PDF.
            # elif '.pdf' in doc and kind != 'url':
            else:
                if verbose:
                    print('Parsing PDF.')
                coa_data = self.parse_pdf(
                    doc,
                    cleanup=cleanup,
                    headers=headers,
                    lims=lims,
                    max_delay=max_delay,
                    persist=persist,
                    resolution=resolution,
                    temp_path=temp_path,
                    use_cached=use_cached,
                    verbose=verbose,
                )

            # TODO: Parse image of COAs and QR codes.
            
            # Record the COA data.
            coas.append(coa_data)

        # Close the parser at the end.
        if persist:
            self.quit()

        # Return the parsed COAs.
        return coas

    def parse_pdf(
            self,
            pdf: Any,
            cleanup: Optional[bool] = True,
            headers: Optional[dict] = {},
            lims: Optional[Any] = None,
            max_delay: Optional[float] = 7,
            persist: Optional[bool] = False,
            resolution: Optional[int] = 300,
            temp_path: Optional[str] = '/tmp',
            use_cached: Optional[bool] = False,
            verbose: Optional[bool] = False,
        ) -> dict:
        """Parse a CoA PDF. Searches the best guess image, then all
        images, for a QR code URL to find results online.
        Args:
            pdf (PDF): A file path to a PDF or a pdfplumber PDF.
            cleanup (bool): Whether or not to remove the files generated
                during OCR, `True` by default (optional).
            headers (dict): Headers for HTTP requests (optional).
            lims (str or dict): Specific LIMS to parse CoAs (optional).
            max_delay (float): The maximum number of seconds to wait
                for the page to load (optional).
            persist (bool): Whether to persist the driver
                and / or session between CoA parses, the
                default is `True`, with any driver and session
                being closed at the end (optional).
            temp_path (str): A temporary directory used for OCR.
            temp_path (str): A temporary directory to store files used
                during PDF OCR, `/tmp` by default (optional).
            resolution (int): The resolution of rendered PDF images,
                300 by default (optional).
        Returns:
            (dict): The sample data.
        """
        if lims is None:
            lims = self.lims

        # Read the PDF.
        if isinstance(pdf, str):
            pdf_file = pdfplumber.open(pdf)
        elif isinstance(pdf, pdfplumber.pdf.PDF):
            pdf_file = pdf
        else:
            with open(magick_wand(file=pdf, resolution=resolution)) as temp_file:
                temp_file.save(f'{temp_path}/coa.pdf')
            pdf_file = pdfplumber.open(f'{temp_path}/coa.pdf')

        # Optional: Try to find any Metrc IDs to query the Cannlytics API.
        # Return cached results instead of parsing the CoA from scratch.
        if use_cached is True:
            metrc_ids = self.find_metrc_ids(pdf_file.pages[0])
            if metrc_ids:
                for metrc_id in metrc_ids:
                    data = self.get_metrc_results(metrc_id)
                    if data is not None:
                        return data

        # Identify any known LIMS, trying OCR if the PDF is not recognized,
        # then raise an error if the labs / LIMS is unknown for safety.
        # TODO: Parse unidentified CoAs to the best of our abilities.
        known_lims = self.identify_lims(pdf_file, lims=lims)
        if verbose:
            print(f'Identified LIMS: {known_lims}')
        if known_lims is None:
            temp_pdf = f'{temp_path}/ocr-coa.pdf'
            if isinstance(pdf, str):
                filename = pdf
            elif isinstance(pdf, pdfplumber.pdf.PDF):
                filename = pdf.stream.name
            else:
                filename = f'{temp_path}/coa.pdf'
            self.pdf_ocr(
                filename,
                temp_pdf,
                temp_path=temp_path,
                resolution=resolution,
                cleanup=cleanup,
            )
            pdf_file = pdfplumber.open(temp_pdf)
            known_lims = self.identify_lims(pdf_file, lims=lims)
            if verbose:
                print(f'Identified LIMS after OCR: {known_lims}')
            if known_lims is None:
                raise ValueError(UNIDENTIFIED_LIMS)

        # Get the time the CoA was created, if known.
        date_tested = self.get_pdf_creation_date(pdf_file)

        # Attempt to use an URL from any QR code on the PDF.
        url = None
        try:
            qr_code_index = self.lims[known_lims].get('qr_code_index')
            url = self.find_pdf_qr_code_url(pdf_file, qr_code_index)
            if url is None and qr_code_index is not None:
                url = self.find_pdf_qr_code_url(pdf_file)
        except IndexError:
            url = self.find_pdf_qr_code_url(pdf_file)
        if verbose:
            print(f'Found URL on PDF: {url}')

        # Get the LIMS parsing routine.
        algorithm_name = LIMS[known_lims]['coa_algorithm_entry_point']
        algorithm = getattr(getattr(self, algorithm_name), algorithm_name)
        if verbose:
            print(f'Using algorithm: {algorithm_name}')

        # Use the URL if found, then try the PDF if the URL fails or is missing.
        if url:
            try:
                if verbose:
                    print(f'Parsing URL: {url}')
                data = self.parse_url(
                    url,
                    headers=headers,
                    lims=known_lims,
                    max_delay=max_delay,
                    persist=persist,
                    verbose=verbose,
                )
            except Exception as e:
                if verbose:
                    print(f'Failed to parse URL:', str(e))
                    print('Trying algorithm instead.')
                data = algorithm(
                    self,
                    pdf_file,
                    headers=headers,
                    max_delay=max_delay,
                    persist=persist,
                    google_maps_api_key=self.google_maps_api_key,
                )
        else:
            if verbose:
                print(f'Parsing PDF: {pdf}')
            data = algorithm(
                self,
                pdf_file,
                headers=headers,
                max_delay=max_delay,
                persist=persist,
                google_maps_api_key=self.google_maps_api_key,
            )

        # Close the PDF.
        pdf_file.close()

        # Return the data.
        sample = {
            'date_tested': date_tested,
            'lab_results_url': url,
            'lims': known_lims,
        }
        return {**sample, **data}

    def parse_url(
            self,
            url: str,
            lims: Optional[Any] = None,
            headers: Optional[dict] = {},
            max_delay: Optional[float] = 7,
            persist: Optional[bool] = False,
            verbose: Optional[bool] = False,
        ) -> dict:
        """Parse a CoA URL.
        Args:
            url (str): The CoA URL.
            lims (str or dict): Specific LIMS to parse CoAs.
            headers (Any): Optional headers for standardization.
            max_delay (float): The maximum number of seconds to wait
                for the page to load.
            persist (bool): Whether to persist the session.
                The default is `False`. If you do persist
                the driver, then make sure to call `quit`
                when you are finished.
        Returns:
            (dict): The sample data.
        """
        if lims is None:
            lims = self.lims

        # Identify the LIMS.
        known_lims = self.identify_lims(url, lims=lims)
        if verbose:
            print(f'Identified LIMS: {known_lims}')

        # Restrict to known labs / LIMS.
        if known_lims is None:
            # raise ValueError(UNIDENTIFIED_LIMS)
        
            # Download URL to a temporary file.
            temp_path = tempfile.gettempdir()
            try:
                filename = url.split('/')[-1].split('?')[0] + '.pdf'
            except:
                filename = 'coa.pdf'
            temp_pdf = os.path.join(temp_path, filename)
            if self.session is not None:
                response = self.session.get(url)
            else:
                response = requests.get(url, headers=headers)
            with open(temp_pdf, 'wb') as pdf:
                pdf.write(response.content)

            # Attempt to identify the LIMS from the PDF.
            known_lims = self.identify_lims(temp_pdf, lims=lims)
            if verbose:
                print(f'Identified LIMS after download: {known_lims}')
            if known_lims is None:
                raise ValueError(UNIDENTIFIED_LIMS)
            else:
                url = temp_pdf

        # Get the LIMS parsing routine.
        algorithm_name = LIMS[known_lims]['coa_algorithm_entry_point']
        if verbose:
            print(f'Using algorithm: {algorithm_name}')
        algorithm = getattr(getattr(self, algorithm_name), algorithm_name)
        data = algorithm(
            self,
            url,
            headers=headers,
            max_delay=max_delay,
            persist=persist,
            google_maps_api_key=self.google_maps_api_key,
        )
        return data

    def pdf_ocr(
            self,
            filename: str,
            outfile: str,
            temp_path: Optional[str] = '/tmp',
            resolution: Optional[int] = 300,
            cleanup: Optional[bool] = True,
        ) -> None:
        """Pass a PDF through OCR to recognize its text. Outputs a new PDF.
        A temporary directory is used, because the algorithm is to:
            1. Convert all PDF pages to images.
            2. Convert each image to PDF with text.
            3. Compile the PDFs with text to a single PDF.
        The rendered images and individual PDF files are removed by default.
        Args:
            filename (str): The filename of the PDF to apply OCR.
            outfile (str): A new PDF file to generate.
            temp_path (str): A temporary directory to store files used
                during PDF OCR, `/tmp` by default (optional).
            resolution (int): The resolution of rendered PDF images,
                300 by default (optional).
            cleanup (bool): Whether or not to remove the files generated
                during OCR, `True` by default (optional).
        """
        # Create a directory to store images and rendered PDFs.
        if not os.path.exists(temp_path): os.makedirs(temp_path)

        # Convert each PDF page to an image.
        image_files = convert_pdf_to_images(
            filename,
            temp_path,
            resolution=resolution,
        )

        # Convert each image to PDF with text.
        pdf_files = []
        for image_file in image_files:
            pdf = image_to_pdf_or_hocr(image_file, extension='pdf')
            pdf_file = image_file.replace('.png', '.pdf')
            pdf_files.append(pdf_file)
            with open(pdf_file, 'w+b') as f:
                f.write(pdf)
                f.close()
            if cleanup:
                os.remove(image_file)

        # Compile the PDFs with text to a single PDF.
        merger = PdfMerger()
        for pdf in pdf_files:
            merger.append(pdf)
        merger.write(outfile)
        merger.close()

        # Remove individual PDF files.
        if cleanup:
            for pdf in pdf_files:
                try:
                    os.remove(pdf)
                except PermissionError:
                    pass

        # Remove all `magick-*` files from the temp directory.
        # for i in os.listdir(temp_path):
        #     path = os.path.join(temp_path, i)
        #     if os.path.isfile(path) and i.startswith('magick-'):
        #         os.remove(path)
    
    def open_pdf_with_ocr(self, doc: str) -> pdfplumber.pdf.PDF:
        """
        Tries to open a PDF document. If an error occurs when opening the PDF,
        then OCR is applied and the PDF is reopened.
        Args:
            parser (Any): An instance of a parser, expected to have a pdf_ocr method.
            doc (str): Path to the PDF document to be parsed.
        Returns:
            pdfplumber.pdf.PDF: An opened PDF document via pdfplumber.
        """
        try:
            report = pdfplumber.open(doc)
        except:
            temp = tempfile.mkstemp('.pdf')[1]
            temp_path = tempfile.gettempdir()
            self.pdf_ocr(doc, temp, temp_path=temp_path)
            try:
                report = pdfplumber.open(temp)
            except:
                raise Exception('Failed to open the PDF even after applying OCR.')
        return report

    def save(
            self,
            data: Any,
            outfile:str,
            codings: Optional[dict] = None,
            column_order: Optional[list] = None,
            nuisance_columns: Optional[list] = None,
            numeric_columns: Optional[list] = None,
            standard_analyses: Optional[dict] = None,
            standard_analytes: Optional[dict] = None,
            standard_fields: Optional[dict] = None,
            google_maps_api_key: Optional[str] = None,
            # TODO: Allow dates to be saved to Excel.
        ) -> Any:
        """Save all CoA data, elongating results and widening values.
        That is, a Workbook is created with a "Details" worksheet that
        has all of the raw data, a "Results" worksheet with long-form
        data where each row is a result for an analyte, and a "Values"
        worksheet with wide-form data where each row is an observation
        and each column is the `value` field for each of the `results`.
        Args:
            data (dict or list or DataFrame): The data to save.
            outfile (str): The file that you wish to save. Accepts a
                response object for returning in an HTTP request.
            codings (dict): A map of value codings, from actual to coding.
            column_order (list): Desired order for columns.
            nuisance_columns (list): A list of column suffixes to remove.
            numeric_columns (list): A list of columns to treat as numeric
                and apply codings.
            standard_analyses (dict): A mapping of encountered analyses to
                standard analyses.
            standard_analytes (dict): A mapping of encountered analytes to
                standard analytes.
            standard_fields (dict): A mapping of encountered fields to
                standard fields.
            google_maps_api_key (str): A Google Maps API Key to supplement
                addresses with latitude and longitude.
        Returns:
            (Workbook): An openpyxl Workbook.
        """
        # Create a workbook for saving the data.
        wb = openpyxl.Workbook()

        # Initialize the details data.
        details_data = None
        if isinstance(data, dict):
            details_data = pd.DataFrame([data])
        elif isinstance(data, list):
            details_data = pd.DataFrame(data)
        else:
            details_data = data

        # Specify the desired order for columns / fields.
        if codings is None:
            codings = self.codings

        # Standardize details.
        details_data = self.standardize(
            details_data,
            column_order=column_order,
            nuisance_columns=nuisance_columns,
            numeric_columns=numeric_columns,
            standard_analyses=standard_analyses,
            standard_analytes=standard_analytes,
            standard_fields=standard_fields,
            google_maps_api_key=google_maps_api_key,
        )

        # Standardize results.
        results_data = self.standardize(
            details_data,
            how='long',
            column_order=column_order,
            nuisance_columns=nuisance_columns,
            numeric_columns=numeric_columns,
            standard_analyses=standard_analyses,
            standard_analytes=standard_analytes,
            standard_fields=standard_fields,
            google_maps_api_key=google_maps_api_key,
        )

        # Standardize values.
        values_data = self.standardize(
            details_data,
            how='wide',
            details_data=details_data,
            results_data=results_data,
            column_order=column_order,
            nuisance_columns=nuisance_columns,
            numeric_columns=numeric_columns,
            standard_analyses=standard_analyses,
            standard_analytes=standard_analytes,
            standard_fields=standard_fields,
            google_maps_api_key=google_maps_api_key,
        )

        # Add a "Codings" worksheet.
        coding_data = pd.DataFrame({
            'Coding': codings.values(),
            'Actual': codings.keys(),
        })
        
        # Format details `results`, `coa_urls`, `images` as proper JSON.
        try:
            details_data['results'] = details_data['results'].apply(json.dumps)
        except:
            pass
        try:
            details_data['coa_urls'] = details_data['coa_urls'].apply(json.dumps)
        except:
            pass
        try:
            details_data['images'] = details_data.get['images'].apply(json.dumps)
        except:
            pass

        # Create a workbook for saving the data.
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = 'Values'
        write_to_worksheet(ws, values_data)
        wb.create_sheet(title='Details')
        ws = wb.worksheets[1]
        write_to_worksheet(ws, details_data)
        wb.create_sheet(title='Results')
        ws = wb.worksheets[2]
        write_to_worksheet(ws, results_data)
        wb.create_sheet(title='Codings')
        ws = wb.worksheets[3]
        write_to_worksheet(ws, coding_data)
        wb.save(outfile)
        wb.close()
        return wb

    def standardize(
            self,
            data: Any,
            codings: Optional[dict] = None,
            column_order: Optional[list] = None,
            nuisance_columns: Optional[list] = None,
            numeric_columns: Optional[list] = None,
            how: Optional[str] = 'details',
            details_data: Optional[Any] = None,
            results_data: Optional[Any] = None,
            standard_analyses: Optional[dict] = None,
            standard_analytes: Optional[dict] = None,
            standard_fields: Optional[dict] = None,
            google_maps_api_key: Optional[str] = None,
        ) -> Any:
        """Standardize (and normalize) given data.
        Args:
            data (dict or list or DataFrame): The data to standardize.
            codings (dict): A map of value codings, from actual to coding.
            column_order (list): A list of columns in desired order.
            nuisance_columns (list): A list of column suffixes to remove.
            numeric_columns (list): A list of columns to treat as numeric
                and apply codings.
            standard_analyses (dict): A mapping of encountered analyses to
                standard analyses.
            standard_analytes (dict): A mapping of encountered analytes to
                standard analytes.
            standard_fields (dict): A mapping of encountered fields to
                standard fields.
            how (str): How to standardize, a simple clean of the data
                `details` by default. Alternatively specify `wide` for a
                wide-form DataFrame of values or `long` for a long-form
                DataFrame of results.
            details_data (DataFrame): The data pre-formatted as `details`.
                May provide a speed increase provided when formatting `values`.
            results_data (DataFrame): The data pre-formatted as `results`.
                May provide a speed increase provided when formatting `values`.
            google_maps_api_key (str): A Google Maps API Key to supplement
                addresses with latitude and longitude.
        Returns:
            (dict or list or DataFrame): Returns the data with standardized
                fields, analyses, analytes, product types, normalized
                results, and augmented with strain name and GIS data.
        """
        # Specify the desired order for columns / fields.
        if codings is None:
            codings = self.codings
        if column_order is None:
            column_order = self.column_order
        if nuisance_columns is None:
            nuisance_columns = self.nuisance_columns
        if numeric_columns is None:
            numeric_columns = self.numeric_columns
        if standard_analyses is None:
            standard_analyses = self.analyses
        if standard_analytes is None:
            standard_analytes = self.analytes
        if standard_fields is None:
            standard_fields = self.fields

        # Future standardization work:
        # [ ] Calculate all missing totals:`total_cannabinoids`, `total_terpenes`, etc.
        # [ ] Remove and keep `units` from `value`.
        # [ ] Standardize `units`
        # [ ] Create a standard `product_type_key`
        # [ ] Augment any missing GIS data, such as latitude, longitude, or address field.
        # [ ] Try to parse a `strain_name` from `product_name` with NLP.

        # Standardize a dictionary.
        if isinstance(data, dict):

            # Identify standard fields, adding analytes for `wide` data.
            fields = standard_fields
            if how == 'wide':
                current_fields = {x:x for x in data.keys()}
                fields = {**current_fields, **standard_analytes}

            # Standardize fields.
            std = {}
            for k, v in data.items():
                key = fields.get(k, k)
                std[key] = v

            # Standardize `analyses` of details.
            if how == 'details':
                std['analyses'] = [standard_analyses.get(x, x) for x in std['analyses']]

                # Standardize the `analysis` of reach of the `results`.
                # Also, normalize the `results`, converting to numeric
                # and applying codings.
                standardized_results = []
                sample_results = data['results']
                if isinstance(sample_results, str):
                    try:
                        sample_results = ast.literal_eval(sample_results)
                        if isinstance(sample_results, str):
                            sample_results = json.load(json.dumps(sample_results))
                    except:
                        json.loads(sample_results)
                for result in sample_results:
                    
                    analysis = result.get('analysis')
                    result['analysis'] = standard_analyses.get(analysis, analysis)
                    for c in numeric_columns:
                        value = result.get(c)
                        value = codings.get(value, value)
                        if isinstance(value, str):
                            value = convert_to_numeric(value, strip=True)
                        result[c] = pd.to_numeric(value, errors='coerce')
                    standardized_results.append(result)
                std['results'] = standardized_results
                
            # Standard `analysis` of long-form data.
            elif how == 'long':
                analysis = std.get('analysis')
                std['analysis'] = standard_analyses.get(analysis, analysis)

            # Normalize the numeric fields for `wide` data.
            if how == 'wide':
                analytes = list([a for a in data.keys() if a not in column_order])
                for c in analytes:
                    value = std.get(c)
                    try:
                        value = codings.get(value, value)
                    except TypeError:
                        pass
                    if isinstance(value, str):
                        value = convert_to_numeric(value, strip=True)
                    std[c] = pd.to_numeric(value, errors='coerce')

            # Normalize numeric fields for `wide` data.
            elif how == 'long':
                for c in numeric_columns:
                    value = std.get(c)
                    value = codings.get(value, value)
                    if isinstance(value, str):
                        value = convert_to_numeric(value, strip=True)
                    std[c] = pd.to_numeric(value, errors='coerce')

            # Turn dates values to ISO format.
            dates = [x for x in std.keys() if x.startswith('date')]
            for k in dates:
                try:
                    std[k] = pd.to_datetime(std[k]).isoformat()
                except:
                    pass

            # Return the standardized observation.
            return std

        # Standardize a list of dictionaries, series, or DataFrames.
        elif isinstance(data, list):
            return [self.standardize(
                x,
                how=how,
                codings=codings,
                column_order=column_order,
                nuisance_columns=nuisance_columns,
                numeric_columns=numeric_columns,
                standard_analyses=standard_analyses,
                standard_analytes=standard_analytes,
                standard_fields=standard_fields,
                google_maps_api_key=google_maps_api_key,
            ) for x in data]

        # Standardize a DataFrame.
        elif isinstance(data, pd.DataFrame):

            # Standardize details (`details` data).
            if how == 'details':

                # Standardize detail columns.
                details_data = data.copy(deep=True)
                details_data.rename(
                    standard_fields,
                    axis=1,
                    inplace=True,
                    errors='ignore',
                )
                details_data = details_data.groupby(level=0, axis=1).first()

                # Drop nuisance columns.
                for c in nuisance_columns:
                    criterion = details_data.columns.str.endswith(c)
                    details_data = details_data.loc[:, ~criterion]

                # Apply codings to results.
                # FIXME: This is super slow (2-3 mins for 2.5k observations)!
                if 'results' in details_data.columns:
                    details_data['results'].replace(codings, inplace=True)
                # except KeyError:
                #     pass

                # Convert totals to numeric.
                # TODO: Calculate totals if they don't already exist:
                totals = [x for x in details_data.keys() if x.startswith('total_')]
                for c in totals:
                    details_data[c] = details_data[c].astype(str).apply(convert_to_numeric, strip=True)
                    details_data[c] = details_data[c].apply(pd.to_numeric, errors='coerce')

                # Re-order columns.
                details_data = reorder_columns(details_data, column_order)
                return details_data            

            # Standardize results (`long` data).
            elif how == 'long':

                # Create a long table of results data.
                results = []
                details_data = data.copy(deep=True)
                for _, item in details_data.iterrows():

                    # Get the sample results.
                    sample_results = item['results']
                    # if isinstance(sample_results, str):
                    try:
                        sample_results = ast.literal_eval(sample_results)
                        if isinstance(sample_results, str):
                            sample_results = json.load(json.dumps(sample_results))
                    except:
                        try:
                            sample_results = json.loads(sample_results)
                        except:
                            pass

                    # Add each entry.
                    for result in sample_results:
                        std = {}
                        for c in column_order:
                            std[c] = item.get(c)
                        try:
                            results.append({**std, **result})
                        except TypeError:
                            results.append({**std, **json.loads(result)})

                # Apply codings (redundant?).
                results_data = pd.DataFrame(results)
                for c in numeric_columns:
                    try:
                        results_data[c] = results_data[c].replace(codings)
                    except KeyError:
                        pass

                # Standardize the results columns.
                results_data.rename(
                    standard_fields,
                    axis=1,
                    inplace=True,
                    errors='ignore',
                )
                results_data = results_data.groupby(level=0, axis=1).first()

                # Standardize the key column (redundant?).
                try:
                    results_data['key'] = results_data['key'].apply(
                        lambda x: ANALYTES.get(x, x)
                    )
                except KeyError:
                    try:
                        results_data['key'] = results_data['name'].apply(
                            lambda x: ANALYTES.get(snake_case(x), snake_case(x))
                        )
                    except KeyError:
                        pass 

                # Re-order columns.
                results_data = reorder_columns(results_data, column_order)
                return results_data

            # Standardize values (`wide` data).
            elif how == 'wide':

                # Get details and results data (if not passed for speed).
                if details_data is None:
                    details_data = self.standardize(
                        data,
                        codings=codings,
                        column_order=column_order,
                        nuisance_columns=nuisance_columns,
                        numeric_columns=numeric_columns,
                        standard_analyses=standard_analyses,
                        standard_analytes=standard_analytes,
                        standard_fields=standard_fields,
                        google_maps_api_key=google_maps_api_key,
                    )
                if results_data is None:
                    results_data = self.standardize(
                        data,
                        how='wide',
                        codings=codings,
                        column_order=column_order,
                        nuisance_columns=nuisance_columns,
                        numeric_columns=numeric_columns,
                        standard_analyses=standard_analyses,
                        standard_analytes=standard_analytes,
                        standard_fields=standard_fields,
                        google_maps_api_key=google_maps_api_key,
                    )

                # Map keys to analysis for ordering for Values worksheet columns.
                # FIXME: Handle observations without results.
                pairs = []
                try:
                    analytes = list(results_data['key'].unique())
                except:
                    analytes = []
                for a in analytes:
                    try:
                        analyses = results_data.loc[results_data['key'] == a]
                        analyses = list(analyses['analysis'].unique())
                        analyses = [x for x in analyses if x is not None]
                        analysis = analyses[0]
                        place = ord(analysis[0])
                        # Hot-fix to order terpenes right after cannabinoids.
                        if place == 116:
                            place = 100 
                        pairs.append((a, analysis, place))
                    except:
                        # Hot-fix to place unidentified analyses at the end.
                        pairs.append((a, None, 122))

                # Sort the pairs of analyses / analytes.
                pairs.sort(key=operator.itemgetter(2))

                # Create a wide table of values data.
                values = []
                for _, item in details_data.iterrows():

                    # Use all of the details in the values.
                    std = item.to_dict()

                    # Old: Only use the default columns. Make optional?
                    # std = {}
                    # for c in column_order:
                    #     std[c] = item[c]

                    # Get the sample results.
                    sample_results = item['results']
                    if isinstance(sample_results, str):
                        try:
                            sample_results = ast.literal_eval(sample_results)
                            if isinstance(sample_results, str):
                                sample_results = json.load(json.dumps(sample_results))
                        except:
                            sample_results = json.loads(sample_results)

                    # Keep the values from each result.
                    for result in sample_results:
                        result = {k: v for k, v in result.items() if v == v}
                        analyte = result.get('key', snake_case(result.get('name')))
                        analyte = standard_analytes.get(analyte, analyte)
                        value = result.get('value', result.get('percent', result.get('mg_g')))
                        std[analyte] = value
                    values.append(std)

                # Rename and combine columns.
                values_data = pd.DataFrame(values)
                values_data.rename(
                    standard_analytes,
                    axis=1,
                    inplace=True,
                    errors='ignore',
                )
                values_data = values_data.groupby(level=0, axis=1).first()

                # Apply codings to columns.
                values_data.replace(codings, inplace=True)

                # Drop nuisance columns.
                for c in nuisance_columns:
                    criterion = values_data.columns.str.endswith(c)
                    values_data = values_data.loc[:, ~criterion]

                # Old: Drop totals. Make optional?
                # criterion = values_data.columns.str.startswith('total_')
                # values_data = values_data.loc[:, ~criterion]

                # Move certain columns to the beginning.
                cols = list(details_data.columns) + [x[0] for x in pairs]
                values_data = reorder_columns(values_data, cols)
                return values_data

        # Standardize a series.
        elif isinstance(data, pd.Series):
            return pd.Series(self.standardize(
                data.to_dict(),
                how=how,
                codings=codings,
                column_order=column_order,
                nuisance_columns=nuisance_columns,
                numeric_columns=numeric_columns,
                standard_analyses=standard_analyses,
                standard_analytes=standard_analytes,
                standard_fields=standard_fields,
                google_maps_api_key=google_maps_api_key,
            ))

        # Raise an error if an incorrect type is passed.
        else:
            raise ValueError
    
    def parse_with_ai(
            self,
            filename: Any,
            temp_path: Optional[str] = None,
            session: Optional[Any] = None,
            headers: Optional[dict] = DEFAULT_HEADERS,
            use_cached: Optional[bool] = False,
            openai_api_key: Optional[str] = None,
            model: Optional[str] = 'gpt-4',
            max_tokens: Optional[int] = 4_000,
            temperature: Optional[float] = 0.0,
            initial_cost: Optional[float] = 0.0,
            instructional_prompt: Optional[str] = None,
            results_prompt: Optional[str] = None,
            coa_prompt: Optional[str] = None,
            max_prompt_length: Optional[int] = 4_000,
            verbose: Optional[bool] = False,
            user: Optional[str] = None,
            retry_pause: Optional[float] = 3.33,
        ) -> list:
        """Parse a COA with AI.
        Args:
            filename (str): The filename or URL of the COA to parse.
        Returns:
            (dict): The parsed CoA data.
        """
        data, prompts, cost = parse_coa_with_ai(
            self,
            filename,
            temp_path=temp_path,
            session=session,
            headers=headers,
            use_cached=use_cached,
            openai_api_key=openai_api_key,
            model=model,
            max_tokens=max_tokens,
            temperature=temperature,
            initial_cost=initial_cost,
            instructional_prompt=instructional_prompt,
            results_prompt=results_prompt,
            coa_prompt=coa_prompt,
            max_prompt_length=max_prompt_length,
            verbose=verbose,
            user=user,
            retry_pause=retry_pause,
        )
        return data, prompts, cost
    
    def scan(
            self,
            filename: Any,
            width: Optional[int] = 1024,
            temp_path: Optional[str] = '/tmp'
        ) -> str:
        """Scan an image for a QR code or barcode and return any data.
        Args:
            filename (str): A path to an image with a barcode or qr code.
            width (int): The base width to resize the image.
        Returns:
            (str): Returns the data from the decoded QR code.
        """
        # Handle the filename.
        if isinstance(filename, str):
            image = cv2.imread(filename)
        elif isinstance(filename, np.ndarray):
            image = filename
        else:
            try:
                image = cv2.imread(filename.filename)
            except:
                raise ValueError('`filename` must be a string or Image.')
        
        # Handle invalid images.
        if image is None:
            raise ValueError('`filename` must be a valid image.')

        # If the temp path has any extension, then use it as the outfile.
        if os.path.splitext(temp_path)[1] != '':
            outfile = temp_path

        # Otherwise, create a temporary file to store the image.
        else:
            if not os.path.exists(temp_path): os.makedirs(temp_path)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
            outfile = os.path.join(temp_path, f'{timestamp}.png')

        # Load the image, apply grayscale, Gaussian blur, and Otsu's threshold.
        # Use morphology to find and connect text contours.
        # Finally, filter for any QR code.
        qr_code_found = False
        original = image.copy()
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (9,9), 0)
        thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5,5))
        close = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)
        contours = cv2.findContours(close, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours = contours[0] if len(contours) == 2 else contours[1]
        for c in contours:
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.04 * peri, True)
            x, y, w, h = cv2.boundingRect(approx)
            area = cv2.contourArea(c)
            ar = w / float(h)
            if len(approx) == 4 and area > 1000 and (ar > .85 and ar < 1.3):
                cv2.rectangle(image, (x, y), (x + w, y + h), (36, 255, 12), 3)
                cv2.imwrite(outfile, original[y: y + h, x: x + w])
                qr_code_found = True

        # If a QR code was not found, then return None.
        if not qr_code_found:
            return None

        # If a width is given, then resize the image to facilitate
        # QR code reading. Calculates the height based on the new width
        # and the original aspect ratio.
        if width:
            img = Image.open(outfile)
            w_percent = (width / float(img.size[0]))
            h_size = int((float(img.size[1]) * float(w_percent)))
            img = img.resize((width, h_size), Image.Resampling.LANCZOS)
            img.save(outfile)

        # Read the resized image again (important) and try to decode QR codes.
        code = None
        image = Image.open(outfile)
        codes = pyzbar.decode(image)
        if codes:
            code = codes[0].data.decode('utf-8')
        return code

    def quit(self):
        """Close any driver, end any session, and reset the parameters."""
        try:
            self.driver.quit()
        except:
            pass
        try:
            self.session.close()
        except:
            pass
        self.driver = None
        self.options = None
        self.session = None
        self.service = None


# === Tests ===
if __name__ == '__main__' and False:

    # Initialize the CoA parser.
    # Future work: Test the parser with different configurations.
    parser = CoADoc()

    # Specify where your data lives for testing.
    DATA_DIR = '../../../.datasets/coas'
    cc_coa_pdf = f'{DATA_DIR}/Classic Jack.pdf'
    cc_coa_url = 'https://share.confidentcannabis.com/samples/public/share/4ee67b54-be74-44e4-bb94-4f44d8294062'
    tagleaf_coa_pdf = f'{DATA_DIR}/Sunbeam.pdf'
    tagleaf_coa_url = 'https://lims.tagleaf.com/coas/F6LHqs9rk9vsvuILcNuH6je4VWCiFzdhgWlV7kAEanIP24qlHS'
    tagleaf_coa_short_url = 'https://lims.tagleaf.com/coa_/F6LHqs9rk9'

    # [✓] TEST: `decode_pdf_qr_code` via `find_pdf_qr_code_url`.
    qr_code_url = parser.find_pdf_qr_code_url(cc_coa_pdf)
    assert qr_code_url.startswith('https')
    qr_code_url = parser.find_pdf_qr_code_url(tagleaf_coa_pdf)
    assert qr_code_url.startswith('https')

    # [✓] TEST: `get_pdf_creation_date`.
    from datetime import datetime
    creation_date = parser.get_pdf_creation_date(cc_coa_pdf)
    assert isinstance(pd.to_datetime(creation_date), datetime)
    creation_date = parser.get_pdf_creation_date(tagleaf_coa_pdf)
    assert isinstance(pd.to_datetime(creation_date), datetime)

    # [✓] TEST: `identify_lims`.
    identified_lims = parser.identify_lims(cc_coa_pdf)
    assert identified_lims == 'Confident Cannabis'
    identified_lims = parser.identify_lims(tagleaf_coa_pdf)
    assert identified_lims == 'TagLeaf LIMS'

    # [✓] TEST: Parse a PDF.
    data = parser.parse_pdf(cc_coa_pdf)

    # [✓] TEST: Parse a URL.
    data = parser.parse_url(cc_coa_url)

    # [✓] TEST: Parse a list of CoA URLs.
    urls = [cc_coa_url, tagleaf_coa_url]
    data = parser.parse(urls)

    # [✓] TEST: Parse a list of CoA PDFs.
    files = [cc_coa_pdf, tagleaf_coa_pdf]
    data = parser.parse(files)

    # [✓] TEST: Parse all CoAs in a given directory.
    data = parser.parse(DATA_DIR)

    # [ ] TEST: Parse all CoAs in a zipped folder!
    zip_folder = '../../../tests/assets/coas/coas.zip'
    data = parser.parse(zip_folder)
    assert data is not None

    # [✓] TEST: Parse a custom CoA (accept an error for now).
    try:
        parser.parse('https://cannlytics.page.link/partial-equilibrium-notes')
    except NotImplementedError:
        pass

    # [✓] TEST: Standardize CoA data.
    coa = 'https://lims.tagleaf.com/coa_/F6LHqs9rk9'
    # coa = '../../../.datasets/coas/Flore COA/Flore Brand/220121OgreInfused.pdf'
    data = parser.parse(coa)

    # [✓] TEST: Standardize CoA data dictionary.
    clean_data = parser.standardize(data[0])

    # [✓] TEST: Standardize CoA data list of dictionaries.
    clean_data_list = parser.standardize(data)

    # [✓] TEST: Standardize CoA data DataFrame.
    dataframe = pd.DataFrame(data)
    details_dataframe = parser.standardize(dataframe)
    results_dataframe = parser.standardize(
        dataframe,
        how='long'
    )
    values_dataframe = parser.standardize(
        dataframe,
        how='wide',
        details_data=details_dataframe,
        results_data=results_dataframe,
    )

    # [✓] TEST: Save CoA data from DataFrame.
    parser.save(dataframe, '../../../tests/assets/coas/test-coas.xlsx')

    # [✓] TEST: Save CoA data from list of dictionaries.
    parser.save(data, '../../../tests/assets/coas/test-coas.xlsx')

    # [✓] TEST: Save CoA data from dictionary.
    parser.save(data[0], '../../../tests/assets/coas/test-coas.xlsx')

    # [✓] TEST: Pass a PDF through OCR.
    doc = '../../../tests/assets/coas/210000068-Cloud-Cake-1g.pdf'
    temp_path = '../../../tests/assets/coas/tmp'
    temp_file = '../../../tests/assets/coas/tmp/ocr-coa.pdf'
    parser.pdf_ocr(doc, temp_file, temp_path=temp_path)

    # [✓] TEST: Parse a COA using OCR using `parse_pdf`.
    doc = '../../../tests/assets/coas/210000068-Cloud-Cake-1g.pdf'
    temp_path = '../../../tests/assets/tmp'
    data = parser.parse_pdf(doc, temp_path=temp_path)

    # [✓] TEST: Parse a COA using OCR using `parse`.
    doc = '../../../tests/assets/coas/210000068-Cloud-Cake-1g.pdf'
    temp_path = '../../../tests/assets/coas/tmp'
    data = parser.parse_pdf(doc, temp_path=temp_path)

    # [ ] TEST: Scan a QR code in an image with the `scan` method.
    parser = CoADoc()
    file_path = '../../../tests/assets/qr-code/steep-hill-qr-code.jpg'
    data = parser.scan(file_path)
    assert data.startswith('https')

    # [✓] TEST: Close the parser.
    parser.quit()
    print('All CoADoc tests finished.')
