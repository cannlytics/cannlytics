"""
Console Views | Cannlytics
Author: Keegan Skeate <keegan@cannlytics.com>
Created: 12/18/2020
Updated: 7/8/2021
"""

# Standard imports
import csv
from json import loads

# External imports
from django.http import HttpResponse
from django.http.response import HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.views.generic.base import TemplateView
import openpyxl

# Internal imports
from cannlytics.firebase import (
    create_log,
    create_session_cookie,
    get_document,
    initialize_firebase,
    update_document,
    revoke_refresh_tokens,
    verify_session_cookie,
    verify_token,
)
from cannlytics.utils.utils import snake_case
from console.state import layout
from console.utils import (
    get_page_data,
    get_page_context,
    get_user_context,
)

BASE = 'console'

#-----------------------------------------------------------------------
# Main view
#-----------------------------------------------------------------------

class ConsoleView(TemplateView):
    """Main view used for most console pages."""

    login_url = '/account/sign-in'
    redirect_field_name = 'redirect_to'

    def get_template_names(self):
        """Get the screen's template based on the URL path, where the
        URL is segmented as 'https://{base}/{screen}/{section}/{unit}.
        A number of page template paths are tried, trying to match a unit
        first, then section, then a screen-section, finally a screen.
        Screen-sections and sections are also search for in a general folder.
        """
        screen = self.kwargs.get('screen', 'dashboard')
        section = self.kwargs.get('section', screen)
        unit = self.kwargs.get('unit', section)
        return [
            f'{BASE}/pages/{screen}/{unit}.html',
            f'{BASE}/pages/{screen}/{section}/{unit}.html',
            f'{BASE}/pages/{screen}/{section}.html',
            f'{BASE}/pages/{screen}/{screen}-{section}.html',
            f'{BASE}/pages/{screen}/{section}/{section}.html',
            f'{BASE}/pages/{screen}/{screen}.html',
            f'{BASE}/pages/misc/{screen}/{screen}-{section}.html',
            f'{BASE}/pages/misc/{screen}/{section}.html',
        ]

    def get_context_data(self, **kwargs):
        """Get context that is used on all pages. The context is retrieved
        dynamically from the app's state. The user's permissions
        are verified on every request. User-specific context and data
        can be returned depending on the page."""
        context = super().get_context_data(**kwargs)
        context['sidebar'] = layout['sidebar']
        context['screen'] = kwargs.get('screen', '')
        context['section'] = kwargs.get('section', '')
        context['unit'] = kwargs.get('unit', '')
        organization_context = context.get('organizations')
        if not context['screen']:
            context['screen'] = 'dashboard'
            context['dashboard'] = layout['dashboard']
        elif organization_context:
            context['organization_context'] = organization_context
        context = get_page_context(self.kwargs, context)
        context = get_page_data(self.kwargs, context)
        context = get_user_context(self.request, context)
        # FIXME: Redirect to the sign in page if there is no user!
        return context


#-----------------------------------------------------------------------
# Authentication views
#-----------------------------------------------------------------------

class LoginView(TemplateView):
    """Dynamic login view for authentication forms."""

    def get_template_names(self):
        page = self.kwargs.get('page', 'login')
        return [f'{BASE}/pages/account/{page}.html']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


def login(request, *args, **argv): #pylint: disable=unused-argument
    """Functional view to create a user session.
    FIXME: Ensure that the request succeeds on the client!
    """
    try:
        authorization = request.headers.get('Authorization', '')
        token = authorization.split(' ').pop()
        if not token:
            return HttpResponse(status=401)
        initialize_firebase()
        session_cookie = create_session_cookie(token)
        response = JsonResponse({"success": True}, status=204)
        response.set_cookie(
            key='__session',
            value=session_cookie,
            # expires=expires, # Optional: Set expiration time.
            # httponly=True, # TODO: Explore httponly option
            # secure=True, # TODO: Explore secure option
        )
        claims = verify_token(token)
        uid = claims['uid']
        create_log(
            ref=f'users/{uid}/logs',
            claims=claims,
            action='Signed in.',
            log_type='auth',
            key='login'
        )
        update_document(f'users/{uid}', {'signed_in': True})
        return response
    except:
        return HttpResponse(status=401)


def logout(request, *args, **argv): #pylint: disable=unused-argument
    """Functional view to remove a user session."""
    try:
        session_cookie = request.COOKIES.get('__session')
        claims = verify_session_cookie(session_cookie)
        uid = claims['uid']
        create_log(
            ref=f'users/{uid}/logs',
            claims=claims,
            action='Signed out.',
            log_type='auth',
            key='logout'
        )
        update_document(f'users/{uid}', {'signed_in': False})
        revoke_refresh_tokens(claims['sub'])
        response = HttpResponse(status=205)
        response.set_cookie('__session', expires=0)
        return response
    except:
        return HttpResponse(status=401)


#-----------------------------------------------------------------------
# Error views
#-----------------------------------------------------------------------

def handler404(request, *args, **argv): #pylint: disable=unused-argument
    """Handle missing pages."""
    template = f'{BASE}/pages/misc/errors/404.html'
    return render(request, template, {}, status=404)


def handler500(request, *args, **argv): #pylint: disable=unused-argument
    """Handle internal errors."""
    template = f'{BASE}/pages/misc/errors/500.html'
    return render(request, template, {}, status=500)


#-----------------------------------------------------------------------
# Helper views
#-----------------------------------------------------------------------

def download_csv_data(request):
    """Download posted data as a CSV file.
    TODO: Pull requested data again (by ID) instead of using posted data.
    TODO: Limit the size / rate of downloads (tie to account usage / billing).
    """
    session_cookie = request.COOKIES.get('__session')
    claims = verify_session_cookie(session_cookie)
    if not claims: 
        return HttpResponse(status=401)
    data = loads(request.body.decode('utf-8'))['data']
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="download.csv"'
    writer = csv.writer(response)
    writer.writerow(list(data[0].keys()))
    for item in data:
        writer.writerow(list(item.values()))
    return response


def import_data(request):
    """Import data from an Excel worksheet for a given data model.
    TODO: Limit the size / rate of downloads (tie to account usage / billing).
    """
    # Authenticate the user.
    session_cookie = request.COOKIES.get('__session')
    claims = verify_session_cookie(session_cookie)
    if not claims: 
        return HttpResponse(status=401)

    # Get the requested import parameters.
    model = request.GET.get('model')
    org_id = request.GET.get('organization_id')
    excel_file = request.FILES['excel_file']
    # Optional: Validations here to check extension and file size.

    # Authorize that the user is part of the organization.
    if org_id not in claims.get('team', []):
        return HttpResponse(status=403)

    # Get singular from data models.
    data_model = get_document(f'organizations/{org_id}/data_models/{model}')
    model_singular = data_model['singular']

    # TODO: Handle .csv imports.

    # Get the worksheet.
    workbook = openpyxl.load_workbook(excel_file)
    sheetnames = workbook.sheetnames
    for sheetname in sheetnames:
        if sheetname == model or sheetname == 'Upload':
            worksheet = workbook[sheetname]

    # Read the imported data.
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)

    # Save imported data to Firestore.
    keys = [snake_case(key) for key in excel_data[0]]
    data = [dict(zip(keys, values)) for values in excel_data[1:]]
    for item in data:
        doc_id = item.get(f'{model_singular}_id', 'uid')
        update_document(f'organizations/{org_id}/{model}/{doc_id}', item)

    # FIXME: Submit form without refresh
    # return JsonResponse({'data': data, 'success': True}, status=200)
    # return HttpResponse(dumps(data), content_type='application/json')
    return HttpResponseRedirect(f'/{model}')


def no_content(request, *args, **argv): #pylint: disable=unused-argument
    """Return an empty response when needed, such as for a ping."""
    return HttpResponse(status=204)
