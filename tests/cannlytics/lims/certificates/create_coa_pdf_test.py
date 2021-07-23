"""
Generate CoAs | Project

Author: Keegan Skeate
Contact: <keegan@cannlytics.com>
Created: 
Updated: 
License: MIT License <https://opensource.org/licenses/MIT>
"""

import openpyxl


def get_worksheet_headers(sheet):
    """Get the headres of a worksheet.
    Args:
        sheet (Worksheet): An openpyx; Excel file object.
    Returns:
        headers (list): A list of header strings.
    """
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    return headers


def get_worksheet_data(sheet, headers):
    """Get the data of a worksheet.
    Args:
        sheet (Worksheet): An openpyx; Excel file object.
        headres (list): A list of headers to map the values.
    Returns:
        list(dict): A list of dictionaries.
    """
    data = []
    for row in sheet.iter_rows(min_row=2):
        values = {}
        for key, cell in zip(headers, row):
            values[key] = cell.value
        data.append(values)
    return data


def read_worksheet(path, sheetname='Upload'):
    """Read the imported data, iterating over the rows and
    getting value from each cell in row.
    Args:
        path (str or InMemoryFile): An Excel workbook to read.
        filename (str): The name of the worksheet to upload.
    Returns:
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.get_sheet_by_name(sheetname)
    headers = get_worksheet_headers(sheet)
    return get_worksheet_data(sheet, headers)


def get_jinja_references(path, sheetname='Template'):
    """Get Jinja-style references in an Excel worksheet."""
    print('Getting Jinja references....')
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[sheetname]
    refs = {}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            value = cell.value
            if not value:
                continue
            elif value.startswith('{{'):
                ref = cell.column_letter + str(cell.row)
                variable = value.replace('{{', '').replace('}}', '').strip()
                refs[variable] = ref

            # TODO: Handle tables
            elif value.startswith('{% for'):
                return NotImplementedError
    return refs


def insert_qr_code():
    """Insert a QR code into a CoA template."""
    print('Inserting QR code...')
    

def create_qr_code():
    """Insert a QR code into a CoA template."""
    print('Inserting QR code...')
    

    
if __name__ == '__main__':
    
    coa_directory = r'C:\Users\keega\Documents\cannlytics\cannlytics\console\static\console\templates\coa_templates'
    coa_template = 'standard_coa_template.xlsx'
    
    
    # Get Jinja references
    filename = coa_directory + '/' + coa_template
    tempalte_refs = get_jinja_references(filename, sheetname='Template')
    
    # Fill in client information
    
    # Dynamically insert all the results.
    
    # Insert signatures
    
    # Insert QR Code
    
    # Create a PDF.
    
    # Upload the PDF.
    
    # Create download link and short link for the CoA.
    
    # Upload the CoA data,
    
    
    
    