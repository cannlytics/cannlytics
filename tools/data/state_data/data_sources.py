"""
Define Data Sources | Cannlytics

Author: Keegan Skeate <keegan@cannlytics.com>  
Created: 8/20/2021  
Updated: 8/20/2021  
License: MIT License <https://opensource.org/licenses/MIT>  

Manage scientific instruments and measurements from the instruments.
"""
from datetime import datetime
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.styles import PatternFill

sources = [
    {
        'state': 'AL',
        'state_name': 'Alabama',
    },
    {
        'state': 'AK',
        'state_name': 'Alaska',
    },
    {
        'state': 'AZ',
        'state_name': 'Arizona',
    },
    {
        'state': 'AR',
        'state_name': 'Arkansas',
    },
    {
        'state': 'CA',
        'state_name': 'California',
    },
    {
        'state': 'CO',
        'state_name': 'Colorado',
    },
    {
        'state': 'CT',
        'state_name': 'Connecticut',
    },
    {
        'state': 'DE',
        'state_name': 'Delaware',
    },
    {
        'state': 'DC',
        'state_name': 'District of Columbia',
    },
    {
        'state': 'FL',
        'state_name': 'Florida',
    },
    {
        'state': 'GA',
        'state_name': 'Georgia',
    },
    {
        'state': 'GU',
        'state_name': 'Guam',
    },
    {
        'state': 'HI',
        'state_name': 'Hawaii',
    },
    {
        'state': 'ID',
        'state_name': 'Idaho',
    },
    {
        'state': 'IL',
        'state_name': 'Illinois',
    },
    {
        'state': 'IN',
        'state_name': 'Indiana',
    },
    {
        'state': 'IA',
        'state_name': 'Iowa',
    },
    {
        'state': 'KS',
        'state_name': 'Kansas',
    },
    {
        'state': 'KY',
        'state_name': 'Kentucky',
    },
    {
        'state': 'LA',
        'state_name': 'Louisiana',
    },
    {
        'state': 'ME',
        'state_name': 'Maine',
    },
    {
        'state': 'MD',
        'state_name': 'Maryland',
    },
    {
        'state': 'MA',
        'state_name': 'Massachusetts',
    },
    {
        'state': 'MI',
        'state_name': 'Michigan',
    },
    {
        'state': 'MN',
        'state_name': 'Minnesota',
    },
    {
        'state': 'MS',
        'state_name': 'Mississippi',
    },
    {
        'state': 'MO',
        'state_name': 'Missouri',
    },
    {
        'state': 'MT',
        'state_name': 'Montana',
    },
    {
        'state': 'NE',
        'state_name': 'Nebraska',
    },
    {
        'state': 'NV',
        'state_name': 'Nevada',
    },
    {
        'state': 'NH',
        'state_name': 'New Hampshire',
    },
    {
        'state': 'NJ',
        'state_name': 'New Jersey',
    },
    {
        'state': 'NM',
        'state_name': 'New Mexico',
    },
    {
        'state': 'NY',
        'state_name': 'New York',
    },
    {
        'state': 'NC',
        'state_name': 'North Carolina',
    },
    {
        'state': 'ND',
        'state_name': 'North Dakota',
    },
    {
        'state': 'OH',
        'state_name': 'Ohio',
    },
    {
        'state': 'OK',
        'state_name': 'Oklahoma',
        'sources': [
            {'name': 'Medical Marijuana Excise Tax', 'url': 'https://oklahomastate.opengov.com/transparency#/33894/accountType=revenues&embed=n&breakdown=types&currentYearAmount=cumulative&currentYearPeriod=months&graph=bar&legendSort=desc&month=5&proration=false&saved_view=105742&selection=A49C34CEBF1D01A1738CB89828C9274D&projections=null&projectionType=null&highlighting=null&highlightingVariance=null&year=2021&selectedDataSetIndex=null&fiscal_start=earliest&fiscal_end=latest'},
            {'name': 'List of Licensed Businesses', 'url': 'https://oklahoma.gov/omma/businesses/list-of-businesses.html'},
        ],
    },
    {
        'state': 'OR',
        'state_name': 'Oregon',
    },
    {
        'state': 'PA',
        'state_name': 'Pennsylvania',
    },
    {
        'state': 'PR',
        'state_name': 'Puerto Rico',
    },
    {
        'state': 'RI',
        'state_name': 'Rhode Island',
    },
    {
        'state': 'SC',
        'state_name': 'South Carolina',
    },
    {
        'state': 'SD',
        'state_name': 'South Dakota',
    },
    {
        'state': 'TN',
        'state_name': 'Tennessee',
    },
    {
        'state': 'TX',
        'state_name': 'Texas',
    },
    {
        'state': 'UT',
        'state_name': 'Utah',
    },
    {
        'state': 'VT',
        'state_name': 'Vermont',
    },
    {
        'state': 'VI',
        'state_name': 'Virgin Islands',
    },
    {
        'state': 'VA',
        'state_name': 'Virginia',
    },
    {
        'state': 'WA',
        'state_name': 'Washington',
    },
    {
        'state': 'WV',
        'state_name': 'West Virginia',
    },
    {
        'state': 'WI',
        'state_name': 'Wisconsin',
    },
    {
        'state': 'WY',
        'state_name': 'Wyoming',
    }
 ]


DATA_POINTS = [
    'state',
    'date',
    'total_sales',
    'total_quantity',
    'total_rec_sales',
    'total_rec_quantity',
    'total_med_sales',
    'total_med_quantity',
    'total_licenses',
    'total_cultivators',
    'total_processors',
    'total_labs',
    'total_retailers',
    'total_transporters',
    'total_patients',
]

def create_data_collection_workbook(sources, date):
    """Create a workbook that is used on a month-by-month
    basis to collect cannabis data for states. The workbook
    can be easily read and aggregated with existing data."""

    # Create workbook and set the sheetname.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'State Data'
    
    # Add the headers and the data.
    sheet.append(DATA_POINTS)
    for source in sources:
        sheet.append([source['state'], date])

    # Style the workbook
    cell_color = Color(rgb='00FFF2CC')
    cell_fill = PatternFill(patternType='solid', fgColor=cell_color)
    bottom_border = Border(bottom=Side(border_style='thin'))
    for i in range(len(DATA_POINTS)):
        sheet.cell(row=1, column=i+1).border = bottom_border
        sheet.cell(row=1, column=i+1).fill = cell_fill

    # Save the workbook.
    directory = './.datasets/monthly_state_data'
    workbook.save(filename=f'{directory}/state_data_{date}.xlsx')


def monthly_state_report():
    """
    Once a month, collect cannabis data points from all states
    with permitted recreational and or medicinal cannabis sales.
    All available monthly series will be collected, with an emphasis
    on the following data points:

        - Total cannabis sales (total_sales)
        - Total cannabis sold (total_quantity)
        - Recreational cannabis sales (total_rec_sales)
        - Recreational cannabis soled (total_rec_quantity)
        - Medicinal cannabis sales (total_med_sales)
        - Medicinal cannabis sold (total_med_quantity)
        - Number of licensees (total_licenses)
        - Number of cultivation licenses (total_cultivators)
        - Number of processor licenses (total_processors)
        - Number of laboratory licenses (total_labs)
        - Number of retail licenses (total_retailers)
        - Number of transportation licenses (total_transporters)

    Optional data points include:

        - Number of medical patients (total_patients)

    Detailed information will be collected about each laboratory.
    Sales data can also be grouped into `flower`, `oil`, and `edibles`.

    Once data is collected, summary statistics and visualization are
    prepared on a state-by-state basis and aggregate totals are
    calculated.

    This will provide a monthly "State of the Industry" report that 
    will include 12-month ahead forecasts.
    """

    # TODO: Collect monthly data for each state.


    # TODO: Aggregate the monthly data.


    # TODO: Calculate statistics.


    # TODO: Prepare figures.


    # TODO: Render LaTeX text and tables.


if __name__ == '__main__':
    
    date = '2021-07-31'
    print('Getting data for the latest month:', date)
    
    
    # Create workbook to collect July 2021 data.
    # create_data_collection_workbook(sources, date)

    # TODO: Ensure data is collected for each month.
    
    # TODO: Ensure data is uploaded for each month.
    
    # TODO: Ensure data can be aggregated.
    
    # TODO: Create the report!
